from openpyxl import load_workbook

ORDER_SHEETS = {
    "Forward Orders",
    "Cancelled or Returned Orders",
}

LISTING_SHEET = "Product Listing"

STORAGE_SHEETS = {
    "Daily Ageing",
    "Aging Reversal",
    "Upfront Storage Charges",
}


def detect_blinkit_report(file):
    """
    Detect the Blinkit Excel report based on workbook sheet names.
    """

    file.seek(0)

    workbook = load_workbook(
        file,
        read_only=True,
        data_only=True,
    )

    sheet_names = set(workbook.sheetnames)

    # ---------------------------------------------------------
    # Order Financial Report
    # ---------------------------------------------------------

    if ORDER_SHEETS.issubset(sheet_names):
        return "ORDER_FINANCIAL"

    # ---------------------------------------------------------
    # Product Listing Report
    # ---------------------------------------------------------

    if LISTING_SHEET in sheet_names:
        return "LISTING"

    # ---------------------------------------------------------
    # Storage Charges Report
    # ---------------------------------------------------------

    if STORAGE_SHEETS.issubset(sheet_names):
        return "STORAGE"

    return None
