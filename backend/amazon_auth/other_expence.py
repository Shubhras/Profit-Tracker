import json
from decimal import Decimal
from django.db import models
from django.http import HttpResponse
from django.contrib.auth import get_user_model
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import serializers, status

from .models import OtherExpense

User = get_user_model()


# ==========================================
# SERIALIZER
# ==========================================
def normalize_marketplace_key(mkt_str):
    if not mkt_str:
        return 'all'
    s = str(mkt_str).strip().lower()
    if 'amazon' in s:
        return 'amazon'
    if 'myntra' in s:
        return 'myntra'
    if 'flipkart' in s:
        return 'flipkart'
    if s in ['all', 'all connected marketplaces']:
        return 'all'
    return s.replace('-', '').replace(' ', '').replace('_', '')


class OtherExpenseSerializer(serializers.ModelSerializer):
    applied_to_count = serializers.SerializerMethodField()
    effective_rate = serializers.SerializerMethodField()

    class Meta:
        model = OtherExpense
        fields = [
            'id',
            'user',
            'expense_name',
            'marketplace',
            'cost_value',
            'start_date',
            'end_date',
            'cost_type',
            'split_lump_sum_by',
            'repeat_monthly',
            'status',
            'applied_to_count',
            'effective_rate',
            'created_at',
            'updated_at',
        ]
        read_only_fields = ['id', 'user', 'created_at', 'updated_at', 'applied_to_count', 'effective_rate']

    def get_applied_to_count(self, obj):
        # Fallback metric display
        if obj.cost_type == 'per_order':
            return "0 orders"
        return "0 SKUs"

    def get_effective_rate(self, obj):
        if obj.cost_type == 'per_order':
            return f"₹{obj.cost_value} / order"
        if obj.split_lump_sum_by == 'net_sales':
            return "By net sales"
        if obj.split_lump_sum_by == 'units_sold':
            return "By units sold"

        try:
            from .models import AmazonListingItem
            user = obj.user
            catalog_count = AmazonListingItem.objects.filter(user=user).values('sku').distinct().count()
            if catalog_count > 0:
                per_sku_rate = Decimal(str(obj.cost_value or 0)) / Decimal(catalog_count)
                return f"₹{per_sku_rate:.2f} / SKU (₹{obj.cost_value} Lump Sum)"
        except Exception:
            pass

        return f"₹{obj.cost_value} Lump Sum"

    def validate(self, attrs):
        request = self.context.get('request')
        user = request.user if request and hasattr(request, 'user') else None
        marketplace = attrs.get('marketplace') or (self.instance.marketplace if self.instance else None)
        cost_type = attrs.get('cost_type') or (self.instance.cost_type if self.instance else None)

        if user and hasattr(user, 'is_authenticated') and user.is_authenticated and marketplace and cost_type:
            target_norm = normalize_marketplace_key(marketplace)
            all_user_exps = OtherExpense.objects.filter(user=user)
            if self.instance and self.instance.pk:
                all_user_exps = all_user_exps.exclude(pk=self.instance.pk)

            matching_exps = [
                e for e in all_user_exps
                if normalize_marketplace_key(e.marketplace) == target_norm
                or target_norm == 'all'
                or normalize_marketplace_key(e.marketplace) == 'all'
            ]

            if matching_exps:
                first_existing = matching_exps[0]
                if first_existing.cost_type != cost_type:
                    existing_type_display = "Per SKU" if first_existing.cost_type == 'per_sku' else "Per Order"
                    new_type_display = "Per SKU" if cost_type == 'per_sku' else "Per Order"
                    raise serializers.ValidationError({
                        'cost_type': f"Marketplace '{marketplace}' is locked to '{existing_type_display}' based on existing '{first_existing.expense_name}' expense. You cannot create a '{new_type_display}' expense for this marketplace."
                    })

        return attrs


# ==========================================
# API VIEWS
# ==========================================
class OtherExpenseListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        marketplace = request.query_params.get('marketplace')
        status_filter = request.query_params.get('status')
        month = request.query_params.get('month') # e.g. '2026-07'

        qs = OtherExpense.objects.filter(user=user)

        if marketplace and marketplace.lower() != 'all':
            qs = qs.filter(marketplace__iexact=marketplace)

        if status_filter:
            qs = qs.filter(status=status_filter)

        if month:
            try:
                year, m = map(int, month.split('-'))
                qs = qs.filter(start_date__year=year, start_date__month=m)
            except Exception:
                pass

        serializer = OtherExpenseSerializer(qs, many=True, context={'request': request})
        return Response({
            'success': True,
            'count': qs.count(),
            'results': serializer.data
        }, status=status.HTTP_200_OK)

    def post(self, request):
        user = request.user
        serializer = OtherExpenseSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            expense = serializer.save(user=user)
            return Response({
                'success': True,
                'message': 'Business expense created successfully',
                'data': OtherExpenseSerializer(expense, context={'request': request}).data
            }, status=status.HTTP_201_CREATED)
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


class OtherExpenseDetailAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk, user):
        try:
            return OtherExpense.objects.get(pk=pk, user=user)
        except OtherExpense.DoesNotExist:
            return None

    def get(self, request, pk):
        expense = self.get_object(pk, request.user)
        if not expense:
            return Response({'success': False, 'message': 'Expense not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = OtherExpenseSerializer(expense, context={'request': request})
        return Response({'success': True, 'data': serializer.data}, status=status.HTTP_200_OK)

    def put(self, request, pk):
        expense = self.get_object(pk, request.user)
        if not expense:
            return Response({'success': False, 'message': 'Expense not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = OtherExpenseSerializer(expense, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            updated_expense = serializer.save()
            return Response({'success': True, 'message': 'Expense updated successfully', 'data': OtherExpenseSerializer(updated_expense, context={'request': request}).data}, status=status.HTTP_200_OK)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        expense = self.get_object(pk, request.user)
        if not expense:
            return Response({'success': False, 'message': 'Expense not found'}, status=status.HTTP_404_NOT_FOUND)
        expense.delete()
        return Response({'success': True, 'message': 'Expense deleted successfully'}, status=status.HTTP_200_OK)


class OtherExpensePreviewAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            cost_value = Decimal(str(request.data.get('cost_value', '0')))
        except Exception:
            cost_value = Decimal('0.00')

        cost_type = request.data.get('cost_type', 'per_sku')
        split_lump_sum_by = request.data.get('split_lump_sum_by', 'equally')

        # Sample preview data
        sku1_units = 120
        sku2_units = 180
        total_units = sku1_units + sku2_units

        if cost_type == 'per_order':
            # Lump sum cost_value distributed per order/unit
            per_order_rate = (cost_value / Decimal(total_units)) if total_units > 0 else Decimal('0')
            sku1_expense = per_order_rate * Decimal(sku1_units)
            sku1_per_unit = per_order_rate
            sku2_expense = per_order_rate * Decimal(sku2_units)
            sku2_per_unit = per_order_rate

            sku1_share_pct = round((sku1_units / total_units) * 100, 1) if total_units > 0 else 0
            sku2_share_pct = round((sku2_units / total_units) * 100, 1) if total_units > 0 else 0

            preview_items = [
                {
                    'sku': 'SAMPLE-SKU-001',
                    'units_sold': sku1_units,
                    'share': f"{sku1_share_pct}%",
                    'expense': float(sku1_expense),
                    'per_unit': float(sku1_per_unit)
                },
                {
                    'sku': 'SAMPLE-SKU-002',
                    'units_sold': sku2_units,
                    'share': f"{sku2_share_pct}%",
                    'expense': float(sku2_expense),
                    'per_unit': float(sku2_per_unit)
                }
            ]
        else:
            # Per SKU lump-sum split
            if split_lump_sum_by == 'equally':
                sku1_share_pct = 50.0
                sku2_share_pct = 50.0
                sku1_expense = cost_value * Decimal('0.50')
                sku2_expense = cost_value * Decimal('0.50')
            elif split_lump_sum_by == 'net_sales':
                sku1_share_pct = 45.0
                sku2_share_pct = 55.0
                sku1_expense = cost_value * Decimal('0.45')
                sku2_expense = cost_value * Decimal('0.55')
            else:  # units_sold
                sku1_share_pct = round((sku1_units / total_units) * 100, 1)
                sku2_share_pct = round((sku2_units / total_units) * 100, 1)
                sku1_expense = cost_value * (Decimal(sku1_units) / Decimal(total_units))
                sku2_expense = cost_value * (Decimal(sku2_units) / Decimal(total_units))

            sku1_per_unit = (sku1_expense / Decimal(sku1_units)) if sku1_units > 0 else Decimal(0)
            sku2_per_unit = (sku2_expense / Decimal(sku2_units)) if sku2_units > 0 else Decimal(0)

            preview_items = [
                {
                    'sku': 'SAMPLE-SKU-001',
                    'units_sold': sku1_units,
                    'share': f"{sku1_share_pct}%",
                    'expense': float(sku1_expense),
                    'per_unit': float(sku1_per_unit)
                },
                {
                    'sku': 'SAMPLE-SKU-002',
                    'units_sold': sku2_units,
                    'share': f"{sku2_share_pct}%",
                    'expense': float(sku2_expense),
                    'per_unit': float(sku2_per_unit)
                }
            ]

        return Response({
            'success': True,
            'cost_value': float(cost_value),
            'cost_type': cost_type,
            'split_lump_sum_by': split_lump_sum_by,
            'preview': preview_items
        }, status=status.HTTP_200_OK)


class OtherExpenseDownloadSampleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import io
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Business Expenses Sample"

        headers = [
            "Expense Name",
            "Marketplace",
            "Cost Value",
            "Cost Type",
            "Split Mode",
            "Start Date",
            "End Date",
            "Repeat Monthly",
            "Status",
        ]
        ws.append(headers)

        sample_rows = [
            ["Storage Expense", "Amazon", 15000, "per_order", "equally", "2026-09-01", "2026-09-30", "Yes", "applied"],
            ["Packaging Box Fee", "Amazon", 2500, "per_sku", "equally", "2026-09-01", "2026-09-30", "No", "applied"],
            ["Software Subscription", "Myntra", 5000, "per_sku", "net_sales", "2026-09-01", "2026-09-30", "Yes", "applied"],
            ["Agency Marketing Fee", "Flipkart", 8000, "per_sku", "units_sold", "2026-09-01", "2026-09-30", "Yes", "draft"],
        ]

        for row in sample_rows:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = 'attachment; filename="Business_Expenses_Sample_Template.xlsx"'
        return response


class OtherExpenseUploadExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import openpyxl
        from datetime import datetime

        user = request.user
        if 'file' not in request.FILES:
            return Response({'success': False, 'message': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        file_obj = request.FILES['file']
        file_name = file_obj.name.lower()

        try:
            if file_name.endswith('.csv'):
                import csv
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))

            if not rows or len(rows) < 2:
                return Response({'success': False, 'message': 'File is empty or missing data rows'}, status=status.HTTP_400_BAD_REQUEST)

            data_rows = rows[1:]
            created_count = 0

            for row in data_rows:
                if not row or not any(row):
                    continue

                expense_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
                marketplace = str(row[1]).strip() if len(row) > 1 and row[1] is not None else 'Amazon'
                try:
                    cost_value = Decimal(str(row[2])) if len(row) > 2 and row[2] not in ['', None] else Decimal('0.00')
                except Exception:
                    cost_value = Decimal('0.00')

                cost_type = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else 'per_sku'
                if cost_type not in ['per_sku', 'per_order']:
                    cost_type = 'per_sku'

                split_mode = str(row[4]).strip().lower() if len(row) > 4 and row[4] is not None else 'equally'
                if split_mode not in ['equally', 'units_sold', 'net_sales']:
                    split_mode = 'equally'

                s_date_str = str(row[5]).strip() if len(row) > 5 and row[5] not in ['', None] else None
                e_date_str = str(row[6]).strip() if len(row) > 6 and row[6] not in ['', None] else None

                start_date = None
                end_date = None
                if s_date_str:
                    try:
                        start_date = datetime.strptime(s_date_str[:10], '%Y-%m-%d').date()
                    except Exception:
                        pass
                if e_date_str:
                    try:
                        end_date = datetime.strptime(e_date_str[:10], '%Y-%m-%d').date()
                    except Exception:
                        pass

                repeat_str = str(row[7]).strip().lower() if len(row) > 7 and row[7] is not None else 'no'
                repeat_monthly = repeat_str in ['yes', 'true', '1']

                status_val = str(row[8]).strip().lower() if len(row) > 8 and row[8] is not None else 'applied'
                if status_val not in ['applied', 'draft']:
                    status_val = 'applied'

                if not expense_name:
                    continue

                existing_exp = OtherExpense.objects.filter(user=user, marketplace__iexact=marketplace).first()
                if existing_exp and existing_exp.cost_type != cost_type:
                    cost_type = existing_exp.cost_type

                OtherExpense.objects.create(
                    user=user,
                    expense_name=expense_name,
                    marketplace=marketplace,
                    cost_value=cost_value,
                    cost_type=cost_type,
                    split_lump_sum_by=split_mode,
                    start_date=start_date,
                    end_date=end_date,
                    repeat_monthly=repeat_monthly,
                    status=status_val,
                )
                created_count += 1

            return Response({
                'success': True,
                'message': f'Successfully imported {created_count} business expenses',
                'count': created_count,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'success': False, 'message': f'Error parsing file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


def calculate_other_expenses_map(user, from_date_local=None, to_date_local=None, items=None, return_total_expense=False):
    """
    Calculates allocated Other Expenses for a list of profitability items/rows.
    
    Each item in `items` is a dict with:
      - 'key': identifier (e.g. parent_asin or sku)
      - 'marketplace': channel/marketplace string (e.g. 'Amazon-India', 'Myntra', 'Amazon')
      - 'units': units sold (int or float or Decimal)
      - 'net_sales': net sales amount (Decimal or float)
      
    Returns a dict: { item_key: Decimal(allocated_expense) }
    or (expense_map, total_effective_expense) if return_total_expense is True.
    """
    from decimal import Decimal
    from django.db.models import Q, Sum
    from .models import OtherExpense, AmazonListingItem, ProductMapping, OrderItem
    try:
        from myntra.models import MyntraListing
    except ImportError:
        MyntraListing = None

    expense_map = {}
    total_effective_expense = Decimal('0')
    if not user or not items:
        if return_total_expense:
            return expense_map, total_effective_expense
        return expense_map

    qs = OtherExpense.objects.filter(user=user).filter(Q(status='applied') | Q(status__isnull=True))

    if from_date_local:
        qs = qs.filter(Q(end_date__gte=from_date_local) | Q(end_date__isnull=True) | Q(repeat_monthly=True))
    if to_date_local:
        qs = qs.filter(Q(start_date__lte=to_date_local) | Q(start_date__isnull=True))

    expenses = list(qs)
    if not expenses:
        if return_total_expense:
            return expense_map, total_effective_expense
        return expense_map

    for exp in expenses:
        exp_mkt = (exp.marketplace or '').strip().lower()
        cost_val = Decimal(str(exp.cost_value or 0))
        if cost_val <= 0:
            continue

        # Filter items matching this expense's marketplace
        matching_items = []
        for item in items:
            item_mkt = (item.get('marketplace') or item.get('channel') or '').strip().lower()
            if (
                exp_mkt in ['all', 'all connected marketplaces', '']
                or exp_mkt in item_mkt
                or item_mkt in exp_mkt
                or ('amazon' in exp_mkt and 'amazon' in item_mkt)
                or ('myntra' in exp_mkt and 'myntra' in item_mkt)
            ):
                matching_items.append(item)

        if not matching_items:
            continue

        # Calculate date proration for the expense over the requested date filter
        effective_cost = cost_val
        if exp.repeat_monthly:
            import calendar
            from datetime import datetime as dt_cls
            filt_start = from_date_local if from_date_local else (exp.start_date or dt_cls.now().date())
            filt_end = to_date_local if to_date_local else (exp.end_date or filt_start)
            
            eff_start = max(filt_start, exp.start_date) if exp.start_date else filt_start
            eff_end = filt_end
            if exp.end_date and exp.start_date:
                if (exp.end_date.year, exp.end_date.month) > (exp.start_date.year, exp.start_date.month):
                    eff_end = min(filt_end, exp.end_date)
            
            if eff_start > eff_end:
                effective_cost = Decimal('0')
            else:
                tot = Decimal('0')
                curr_year = eff_start.year
                curr_month = eff_start.month
                
                while True:
                    m_start = dt_cls(curr_year, curr_month, 1).date()
                    last_day = calendar.monthrange(curr_year, curr_month)[1]
                    m_end = dt_cls(curr_year, curr_month, last_day).date()
                    
                    ov_start = max(eff_start, m_start)
                    ov_end = min(eff_end, m_end)
                    
                    if ov_start <= ov_end:
                        ov_days = (ov_end - ov_start).days + 1
                        m_days = last_day
                        tot += cost_val * Decimal(str(ov_days)) / Decimal(str(m_days))
                    
                    if curr_year == eff_end.year and curr_month == eff_end.month:
                        break
                    
                    curr_month += 1
                    if curr_month > 12:
                        curr_month = 1
                        curr_year += 1
                
                effective_cost = tot
        elif exp.start_date and exp.end_date:
            exp_start = exp.start_date
            exp_end = exp.end_date
            exp_total_days = (exp_end - exp_start).days + 1
            if exp_total_days > 0:
                filt_start = from_date_local if from_date_local else exp_start
                filt_end = to_date_local if to_date_local else exp_end
                
                overlap_start = max(exp_start, filt_start)
                overlap_end = min(exp_end, filt_end)
                
                if overlap_start <= overlap_end:
                    overlap_days = (overlap_end - overlap_start).days + 1
                    effective_cost = cost_val * Decimal(str(overlap_days)) / Decimal(str(exp_total_days))
                else:
                    effective_cost = Decimal('0')

        if effective_cost <= 0:
            continue

        total_effective_expense += effective_cost

        if exp.cost_type == 'per_order':
            # Per Order: effective_cost distributed per order/unit sold across catalog orders in the period.
            catalog_units = None
            try:
                if 'myntra' in exp_mkt:
                    from myntra.models import MyntraOrder
                    mo_qs = MyntraOrder.objects.filter(myntra_connection__user=user)
                    if from_date_local:
                        mo_qs = mo_qs.filter(created_on__date__gte=from_date_local)
                    if to_date_local:
                        mo_qs = mo_qs.filter(created_on__date__lte=to_date_local)
                    m_cnt = mo_qs.count()
                    if m_cnt > 0:
                        catalog_units = Decimal(str(m_cnt))
                else:
                    oi_qs = OrderItem.objects.filter(order__user=user)
                    if from_date_local:
                        oi_qs = oi_qs.filter(order__purchase_date__date__gte=from_date_local)
                    if to_date_local:
                        oi_qs = oi_qs.filter(order__purchase_date__date__lte=to_date_local)
                    s_qty = oi_qs.aggregate(s=Sum('quantity_ordered'))['s']
                    if s_qty and s_qty > 0:
                        catalog_units = Decimal(str(s_qty))
            except Exception:
                catalog_units = None

            if not catalog_units or catalog_units <= 0:
                catalog_units = sum(Decimal(str(item.get('units') or 0)) for item in matching_items)

            if not catalog_units or catalog_units <= 0:
                catalog_units = Decimal(str(len(matching_items) or 1))

            unit_cost_per_order = effective_cost / catalog_units

            for item in matching_items:
                key = item['key']
                units = Decimal(str(item.get('units') or 0))
                allocated = unit_cost_per_order * units
                expense_map[key] = expense_map.get(key, Decimal('0')) + allocated
        else:
            # Per SKU: Distribute effective_cost equally across active SKUs in the user's account for the requested date range.
            parent_sku_map = {}

            if 'amazon' in exp_mkt or exp_mkt in ['all', 'all connected marketplaces', '']:
                active_parents = set()

                oi_qs = OrderItem.objects.filter(order__user=user)
                if from_date_local:
                    oi_qs = oi_qs.filter(order__purchase_date__date__gte=from_date_local)
                if to_date_local:
                    oi_qs = oi_qs.filter(order__purchase_date__date__lte=to_date_local)

                for row in oi_qs.values('parent_asin', 'asin'):
                    p = row.get('parent_asin') or row.get('asin')
                    if p:
                        active_parents.add(p)

                try:
                    from amazon_ads.models import ProductAdMetric
                    ad_qs = ProductAdMetric.objects.filter(product_ad__amazon_account__user=user, cost__gt=0)
                    if from_date_local:
                        ad_qs = ad_qs.filter(report_date__gte=from_date_local)
                    if to_date_local:
                        ad_qs = ad_qs.filter(report_date__lte=to_date_local)
                    ad_skus = set(ad_qs.values_list('product_ad__sku', flat=True))

                    pm_dict = {m['seller_sku']: (m.get('parent_asin') or m.get('asin')) for m in ProductMapping.objects.filter(account__user=user, seller_sku__in=ad_skus).values('seller_sku', 'parent_asin', 'asin')}
                    ali_dict = {m['sku']: m.get('asin') for m in AmazonListingItem.objects.filter(user=user, sku__in=ad_skus).values('sku', 'asin')}

                    for s in ad_skus:
                        p = pm_dict.get(s) or ali_dict.get(s) or s
                        if p:
                            active_parents.add(p)
                except Exception:
                    pass

                parent_sku_map = {p: set() for p in active_parents}

                for row in oi_qs.values('parent_asin', 'asin', 'seller_sku'):
                    p = row.get('parent_asin') or row.get('asin')
                    s = row.get('seller_sku')
                    if p in parent_sku_map and s:
                        parent_sku_map[p].add(s)

                pm_qs = ProductMapping.objects.filter(account__user=user, parent_asin__in=active_parents).values('parent_asin', 'seller_sku')
                for pm in pm_qs:
                    p = pm.get('parent_asin')
                    s = pm.get('seller_sku')
                    if p in parent_sku_map and s:
                        parent_sku_map[p].add(s)

                ali_qs = AmazonListingItem.objects.filter(user=user, asin__in=active_parents).values('asin', 'sku')
                for ali in ali_qs:
                    p = ali.get('asin')
                    s = ali.get('sku')
                    if p in parent_sku_map and s:
                        parent_sku_map[p].add(s)

            if 'myntra' in exp_mkt or exp_mkt in ['all', 'all connected marketplaces', '']:
                try:
                    from myntra.models import MyntraOrder
                    mo_qs = MyntraOrder.objects.filter(myntra_connection__user=user)
                    if from_date_local:
                        mo_qs = mo_qs.filter(created_on__date__gte=from_date_local)
                    if to_date_local:
                        mo_qs = mo_qs.filter(created_on__date__lte=to_date_local)
                    for s in mo_qs.values_list('seller_sku_code', flat=True).distinct():
                        if s:
                            parent_sku_map.setdefault(s, set()).add(s)
                except Exception:
                    pass

            total_account_skus_cnt = Decimal(str(sum(len(skus) for skus in parent_sku_map.values()))) if parent_sku_map else Decimal('0')
            report_skus_cnt = sum(
                (Decimal(str(item.get('sku_count') or 1)) / (Decimal(str(item.get('order_count_for_sku') or 1)) if Decimal(str(item.get('order_count_for_sku') or 1)) > 0 else Decimal('1')))
                for item in matching_items
            )
            total_active_skus_cnt = max(report_skus_cnt, total_account_skus_cnt)
            if total_active_skus_cnt <= 0:
                total_active_skus_cnt = Decimal(str(len(matching_items) or 1))

            unit_cost_per_sku = effective_cost / total_active_skus_cnt

            for item in matching_items:
                key = item['key']
                sku_cnt = Decimal(str(item.get('sku_count') or 1))
                ord_cnt = Decimal(str(item.get('order_count_for_sku') or 1))
                if ord_cnt <= 0:
                    ord_cnt = Decimal(1)
                allocated = (unit_cost_per_sku * sku_cnt) / ord_cnt
                expense_map[key] = expense_map.get(key, Decimal('0')) + allocated

    if return_total_expense:
        return expense_map, total_effective_expense
    return expense_map



