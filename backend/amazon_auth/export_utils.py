import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

def generate_csv(data_list, headers, keys, totals_dict=None) -> bytes:
    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(headers)
    
    for row in data_list:
        writer.writerow([row.get(k, '') for k in keys])
        
    if totals_dict:
        total_row = []
        for i, k in enumerate(keys):
            if i == 0:
                total_row.append("Total")
            else:
                total_row.append(totals_dict.get(k, ''))
        writer.writerow(total_row)
        
    return stream.getvalue().encode('utf-8')

def generate_xlsx(data_list, headers, keys, totals_dict=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    totals_font = Font(name="Calibri", size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    totals_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # Write headers
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write data
    for row in data_list:
        row_data = [row.get(k, '') for k in keys]
        cleaned_row = []
        for val in row_data:
            if isinstance(val, (list, dict)):
                cleaned_row.append(str(val))
            elif val is None:
                cleaned_row.append('')
            else:
                cleaned_row.append(val)
        ws.append(cleaned_row)
        curr_row = ws.max_row
        for col_idx in range(1, len(keys) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
    # Write totals
    if totals_dict:
        total_row = []
        for i, k in enumerate(keys):
            if i == 0:
                total_row.append("Total")
            else:
                total_row.append(totals_dict.get(k, ''))
        ws.append(total_row)
        curr_row = ws.max_row
        for col_idx in range(1, len(keys) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = totals_font
            cell.border = totals_border
            
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val) + 4)
            else:
                max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)
        
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

def generate_pdf(data_list, headers, keys, filename, totals_dict=None) -> bytes:
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
        textColor=colors.HexColor('#1F4E78'),
        alignment=1,
        spaceAfter=10
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        fontName='Helvetica-Bold',
        fontSize=5.5,
        leading=6.5,
        textColor=colors.white,
        alignment=1,
    )
    
    cell_style = ParagraphStyle(
        'TableCell',
        fontName='Helvetica',
        fontSize=5,
        leading=6,
        textColor=colors.black,
        alignment=0,
    )
    
    totals_style = ParagraphStyle(
        'TableTotals',
        fontName='Helvetica-Bold',
        fontSize=5,
        leading=6,
        textColor=colors.black,
        alignment=0,
    )
    
    elements = []
    elements.append(Paragraph(filename.replace('_', ' ').title(), title_style))
    elements.append(Spacer(1, 8))
    
    table_data = []
    header_row = [Paragraph(str(h), header_style) for h in headers]
    table_data.append(header_row)
    
    for row in data_list:
        data_row = []
        for k in keys:
            val = row.get(k, '')
            if isinstance(val, (list, dict)):
                val = str(val)
            elif val is None:
                val = ''
            data_row.append(Paragraph(str(val), cell_style))
        table_data.append(data_row)
        
    if totals_dict:
        totals_row = []
        for i, k in enumerate(keys):
            if i == 0:
                totals_row.append(Paragraph("Total", totals_style))
            else:
                val = totals_dict.get(k, '')
                totals_row.append(Paragraph(str(val), totals_style))
        table_data.append(totals_row)
        
    col_count = len(keys)
    if col_count > 0:
        col_widths = [805.0 / col_count] * col_count
    else:
        col_widths = [805.0]
        
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 3),
        ('TOPPADDING', (0, 0), (-1, 0), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2),
        ('TOPPADDING', (0, 1), (-1, -1), 2),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
    ]
    
    if totals_dict:
        t_style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F2F2F2')))
        t_style.append(('LINEABOVE', (0, -1), (-1, -1), 1.0, colors.black))
        t_style.append(('LINEBELOW', (0, -1), (-1, -1), 1.5, colors.black))
        
    t.setStyle(TableStyle(t_style))
    elements.append(t)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf
