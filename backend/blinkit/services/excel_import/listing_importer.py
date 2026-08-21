from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from blinkit.models import BlinkitOrderItem, BlinkitProduct

LISTING_SHEET = "Product Listing"

# Actual Blinkit Listing Report structure
HEADER_ROW = 3
DATA_START_ROW = 4


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------


def clean_value(value):
    """
    Clean Excel values.
    """

    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if not value or value in {"-", "--", "NA", "N/A"}:
            return None

        return value

    return value


def normalize_text(value):
    value = clean_value(value)

    if value is None:
        return None

    return str(value).strip()


def normalize_id(value):
    """
    Convert IDs to consistent strings.

    Example:
        10327165
        10327165.0

    both become:

        "10327165"
    """

    value = clean_value(value)

    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def to_decimal_or_none(value):
    """
    Convert Excel numeric values to Decimal.
    """

    value = clean_value(value)

    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    try:
        if isinstance(value, str):
            value = value.replace("₹", "").replace(",", "").replace("%", "").strip()

        return Decimal(str(value))

    except (InvalidOperation, ValueError, TypeError):
        return None


# ---------------------------------------------------------------------
# Header normalization
# ---------------------------------------------------------------------


def normalize_header(value):
    """
    Normalize an Excel header so small formatting differences don't
    break the importer.

    Example:

        "Brand name"
        " Brand name "
        "Brand name\n"

    all become:

        "brand name"
    """

    value = clean_value(value)

    if value is None:
        return None

    return " ".join(str(value).strip().lower().split())


def build_header_map(worksheet):
    """
    Blinkit Listing Report:

        Row 2 -> section titles
        Row 3 -> actual headers

    Returns:

        {
            "product id": 1,
            "brand name": 2,
            ...
        }
    """

    headers = []

    for column_number, cell in enumerate(
        worksheet[HEADER_ROW],
        start=1,
    ):
        header = normalize_header(cell.value)

        if header:
            headers.append((header, column_number))

    if not headers:
        raise ValueError("Product Listing header row is empty.")

    return dict(headers)


# ---------------------------------------------------------------------
# Required columns
# ---------------------------------------------------------------------


def validate_required_columns(header_map):
    """
    Validate only fields that are essential for creating a product.

    Item ID is the most important field because it connects the
    Listing Report with the Order Financial Report.
    """

    required_columns = {
        "product id",
        "brand name",
        "product name",
        "expansion level",
        "uom",
        "mrp",
        "selling price",
        "item id",
        "upc/ean/upc exemption code",
        "parent company (manufacturer)",
    }

    missing_columns = required_columns - set(header_map.keys())

    if missing_columns:
        raise ValueError(
            "Product Listing report is missing required columns: "
            + ", ".join(sorted(missing_columns))
        )


# ---------------------------------------------------------------------
# Row reader
# ---------------------------------------------------------------------


def get_cell_value(row, header_map, column_name):
    """
    Get a value from a row using the normalized column name.
    """

    normalized_column = normalize_header(column_name)

    column_number = header_map.get(normalized_column)

    if column_number is None:
        return None

    # openpyxl rows are zero-indexed when converted to a list.
    index = column_number - 1

    if index >= len(row):
        return None

    return row[index]


def get_listing_rows(worksheet, header_map):
    """
    Read Product Listing rows starting from row 4.
    """

    for row in worksheet.iter_rows(
        min_row=DATA_START_ROW,
        values_only=True,
    ):
        if not any(value is not None for value in row):
            continue

        yield row


# ---------------------------------------------------------------------
# Product import
# ---------------------------------------------------------------------


def import_listing_row(
    *,
    row,
    header_map,
    account,
    import_batch,
):
    """
    Import one Product Listing row.

    Product identity:

        account + item_id

    Existing product:
        UPDATE

    New product:
        CREATE
    """

    # -------------------------------------------------------------
    # Core product information
    # -------------------------------------------------------------

    item_id = normalize_id(
        get_cell_value(
            row,
            header_map,
            "Item ID",
        )
    )

    if not item_id:
        raise ValueError("Product Listing row is missing Item ID.")

    product_id = normalize_id(
        get_cell_value(
            row,
            header_map,
            "Product ID",
        )
    )

    product_name = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Product name",
        )
    )

    brand_name = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Brand name",
        )
    )

    expansion_level = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Expansion level",
        )
    )

    uom = normalize_text(
        get_cell_value(
            row,
            header_map,
            "UoM",
        )
    )

    # -------------------------------------------------------------
    # Pricing
    # -------------------------------------------------------------

    mrp = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "MRP",
        )
    )

    selling_price = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Selling price",
        )
    )

    # -------------------------------------------------------------
    # Identification
    # -------------------------------------------------------------

    upc_ean = normalize_text(
        get_cell_value(
            row,
            header_map,
            "UPC/EAN/UPC Exemption Code",
        )
    )

    manufacturer = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Parent company (Manufacturer)",
        )
    )

    # -------------------------------------------------------------
    # Product classification
    # -------------------------------------------------------------

    hsn = normalize_text(
        get_cell_value(
            row,
            header_map,
            "HSN",
        )
    )

    business_category = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Business category",
        )
    )

    # -------------------------------------------------------------
    # Marketplace fees
    # -------------------------------------------------------------

    commission_percent = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Platform commission (%)",
        )
    )

    commission_value = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Platform commission value",
        )
    )

    fulfillment_fee = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Fulfillment fee",
        )
    )

    inwarding_fee = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Inwarding fee",
        )
    )

    # -------------------------------------------------------------
    # Save product
    # -------------------------------------------------------------

    defaults = {
        "product_id": product_id,
        "product_name": product_name,
        "brand_name": brand_name,
        "upc_ean": upc_ean,
        "mrp": mrp,
        "selling_price": selling_price,
        "hsn": hsn,
        "business_category": business_category,
        "platform_commission_percent": commission_percent,
        "platform_commission_value": commission_value,
        "fulfillment_fee": fulfillment_fee,
        "inwarding_fee": inwarding_fee,
        "uom": uom,
        "expansion_level": expansion_level,
        "manufacturer": manufacturer,
        "source_import": import_batch,
    }

    product, created = BlinkitProduct.objects.update_or_create(
        account=account,
        item_id=item_id,
        defaults=defaults,
    )
    
    BlinkitOrderItem.objects.filter(
        order__account=account,
        item_id=item_id,
        product__isnull=True,
    ).update(
        product=product,
    )

    return {
        "created": created,
        "updated": not created,
        "product": product,
    }


# ---------------------------------------------------------------------
# Main importer
# ---------------------------------------------------------------------


@transaction.atomic
def import_listing_report(
    *,
    file,
    account,
    import_batch,
):
    """
    Import Blinkit Product Listing report.

    Actual workbook structure:

        Sheet:
            Product Listing

        Row 2:
            Product details

        Row 3:
            Headers

        Row 4+:
            Product data
    """

    file.seek(0)

    workbook = load_workbook(
        file,
        read_only=False,
        data_only=True,
    )

    # -------------------------------------------------------------
    # Validate sheet
    # -------------------------------------------------------------

    if LISTING_SHEET not in workbook.sheetnames:
        raise ValueError(f"Missing required sheet: {LISTING_SHEET}")

    worksheet = workbook[LISTING_SHEET]

    # -------------------------------------------------------------
    # Build header map from ROW 3
    # -------------------------------------------------------------

    header_map = build_header_map(worksheet)

    # -------------------------------------------------------------
    # Validate required columns
    # -------------------------------------------------------------

    validate_required_columns(header_map)

    # -------------------------------------------------------------
    # Read rows starting from ROW 4
    # -------------------------------------------------------------

    rows = list(
        get_listing_rows(
            worksheet,
            header_map,
        )
    )

    if not rows:
        raise ValueError("Product Listing report contains no product rows.")

    # -------------------------------------------------------------
    # Import products
    # -------------------------------------------------------------

    created_count = 0
    updated_count = 0
    skipped_count = 0

    for row in rows:
        try:
            result = import_listing_row(
                row=row,
                header_map=header_map,
                account=account,
                import_batch=import_batch,
            )

        except ValueError:
            raise

        if result["created"]:
            created_count += 1

        elif result["updated"]:
            updated_count += 1

        else:
            skipped_count += 1

    # -------------------------------------------------------------
    # Return result
    # -------------------------------------------------------------

    return {
        "total_rows": len(rows),
        "products_created": created_count,
        "products_updated": updated_count,
        "products_skipped": skipped_count,
    }
