from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from blinkit.models import (
    BlinkitOrder,
    BlinkitOrderItem,
    BlinkitPayout,
    BlinkitProduct,
)

FORWARD_SHEET = "Forward Orders"
RETURN_SHEET = "Cancelled or Returned Orders"


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------


def serialize_raw_value(value):

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, Decimal):
        return str(value)

    return value


def serialize_raw_row(row):
    return {str(key): serialize_raw_value(value) for key, value in row.items()}


def clean_value(value):
    """
    Convert Excel empty values / '-' into None.
    """
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if not value or value in {"-", "--", "NA", "N/A"}:
            return None

        return value

    return value


def to_decimal(value, default=Decimal("0")):
    """
    Safely convert Excel numeric values to Decimal.
    """
    value = clean_value(value)

    if value is None:
        return default

    if isinstance(value, Decimal):
        return value

    try:
        if isinstance(value, str):
            value = value.replace("₹", "").replace(",", "").replace("%", "").strip()

        return Decimal(str(value))

    except (InvalidOperation, ValueError, TypeError):
        return default


def to_decimal_or_none(value):
    """
    Same as to_decimal(), but returns None when the Excel cell
    is empty / '-'.
    """
    value = clean_value(value)

    if value is None:
        return None

    return to_decimal(value)


def parse_date(value):
    """
    Parse Blinkit date values.

    Examples from the report:
        1 June 2026
        15 June 2026
        18 June 2026
        2026-06-18 05:30:00
    """

    value = clean_value(value)

    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    value = str(value).strip()

    formats = [
        "%d %B %Y",
        "%d %b %Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue

    raise ValueError(f"Unable to parse Blinkit date: {value}")


def normalize_text(value):
    """
    Convert Excel value to a clean string.
    """
    value = clean_value(value)

    if value is None:
        return None

    return str(value).strip()


def normalize_id(value):
    """
    Excel may give numeric IDs as integers/floats.
    Store them consistently as strings.
    """
    value = clean_value(value)

    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def derive_payout_period(settlement_date):
    """
    Blinkit payouts are based on 15-day cycles.

    1st - 15th
    16th - month end

    Example:
        settlement = 18 June 2026
        payout period = 1 June 2026 -> 15 June 2026
    """

    if not settlement_date:
        raise ValueError("Settlement date is required to derive payout period.")

    settlement_date = settlement_date.date()

    if settlement_date.day <= 15:
        start_date = settlement_date.replace(day=1)
        end_date = settlement_date.replace(day=15)
    else:
        start_date = settlement_date.replace(day=16)

        last_day = monthrange(
            settlement_date.year,
            settlement_date.month,
        )[1]

        end_date = settlement_date.replace(day=last_day)

    return start_date, end_date


# ---------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------


def get_sheet_rows(worksheet, header_row=5, data_start_row=7):
    """
    Return worksheet rows as dictionaries.

    Blinkit reports have title/blank rows before the actual headers.
    """

    headers = [clean_value(cell.value) for cell in worksheet[header_row]]

    if not headers or not headers[0]:
        raise ValueError(f"Unable to find headers in sheet: {worksheet.title}")

    for row in worksheet.iter_rows(
        min_row=data_start_row,
        values_only=True,
    ):
        if not any(value is not None for value in row):
            continue

        yield dict(zip(headers, row))


# ---------------------------------------------------------------------
# Payout
# ---------------------------------------------------------------------


def get_or_create_payout(
    *,
    account,
    settlement_date,
    bank_utr,
    settlement_status,
):
    """
    Get/create the payout represented by the current report row.

    For the current report all rows belong to the same payout.
    The grouping is still done from the actual settlement information
    rather than assuming one payout globally.
    """

    payout_start, payout_end = derive_payout_period(settlement_date)

    payout, created = BlinkitPayout.objects.get_or_create(
        account=account,
        payout_period_start=payout_start,
        payout_period_end=payout_end,
        defaults={
            "bank_utr": bank_utr,
            "settlement_date": settlement_date.date(),
            "settlement_status": settlement_status,
        },
    )

    if not created:
        # Fill missing settlement information if an older import
        # created the payout without complete information.
        changed = False

        if not payout.bank_utr and bank_utr:
            payout.bank_utr = bank_utr
            changed = True

        if not payout.settlement_date and settlement_date:
            payout.settlement_date = settlement_date.date()
            changed = True

        if not payout.settlement_status and settlement_status:
            payout.settlement_status = settlement_status
            changed = True

        if changed:
            payout.save(
                update_fields=[
                    "bank_utr",
                    "settlement_date",
                    "settlement_status",
                    "updated_at",
                ]
            )

    return payout


# ---------------------------------------------------------------------
# Product linking
# ---------------------------------------------------------------------


def find_product(account, item_id):
    """
    Listing data may or may not have been uploaded before the order report.

    Therefore:
        product exists -> link it
        product doesn't exist -> keep product NULL

    The order import must not fail simply because the Listing Report
    has not been uploaded yet.
    """

    if not item_id:
        return None

    return BlinkitProduct.objects.filter(
        account=account,
        item_id=item_id,
    ).first()


# ---------------------------------------------------------------------
# Forward Orders
# ---------------------------------------------------------------------


def import_forward_row(
    *,
    row,
    account,
    import_batch,
    payout,
):
    """
    Import one row from the Forward Orders sheet.
    """

    order_id = normalize_id(row.get("Order ID"))
    invoice_id = normalize_text(row.get("Invoice ID"))
    item_id = normalize_id(row.get("Item ID"))

    if not order_id:
        raise ValueError("Missing Order ID.")

    if not item_id:
        raise ValueError(f"Missing Item ID for Order ID {order_id}.")

    order_date = parse_date(row.get("Order Date"))

    order, order_created = BlinkitOrder.objects.get_or_create(
        account=account,
        order_id=order_id,
        order_kind="FORWARD",
        defaults={
            "invoice_id": invoice_id,
            "order_type": normalize_text(row.get("Order Type")),
            "order_date": order_date,
            "order_status": normalize_text(row.get("Order Status")),
            "customer_name": normalize_text(row.get("Customer Name")),
            "customer_gst_name": normalize_text(
                row.get("GST Name (where buyer has given gst)")
            ),
            "customer_gst_number": normalize_text(
                row.get("GST Number (where buyer has given gst)")
            ),
            "supply_state": normalize_text(row.get("Supply State")),
            "customer_city": normalize_text(row.get("Customer City")),
            "customer_state": normalize_text(row.get("Customer State")),
            # COMPLETE ORIGINAL EXCEL ROW
            "raw_data": serialize_raw_row(row),
            "payout": payout,
            "source_import": import_batch,
        },
    )

    # If the order already existed from a previous upload,
    # don't overwrite its original financial/order data.
    #
    # But make sure payout/source import are populated if missing.
    if not order_created:
        changed = False

        if order.payout_id is None:
            order.payout = payout
            changed = True

        if order.source_import_id is None:
            order.source_import = import_batch
            changed = True

        if changed:
            order.save(
                update_fields=[
                    "payout",
                    "source_import",
                    "updated_at",
                ]
            )

    # -------------------------------------------------------------
    # Duplicate protection
    #
    # Current report has no duplicate (Order ID + Item ID) pairs.
    # We use:
    #
    #     order + item_id
    #
    # as the natural key for an order item.
    # -------------------------------------------------------------

    existing_item = BlinkitOrderItem.objects.filter(
        order=order,
        item_id=item_id,
    ).first()

    if existing_item:
        return {
            "order_created": order_created,
            "item_created": False,
            "skipped": True,
        }

    product = find_product(
        account=account,
        item_id=item_id,
    )

    order_item = BlinkitOrderItem.objects.create(
        order=order,
        product=product,
        item_id=item_id,
        product_name=normalize_text(row.get("Product Name")),
        variant_description=normalize_text(row.get("Variant Description")),
        quantity=to_decimal(row.get("Quantity")),
        mrp=to_decimal_or_none(row.get("MRP (Rs)")),
        selling_price=to_decimal_or_none(row.get("Selling Price (Rs)")),
        igst_percent=to_decimal_or_none(row.get("IGST %")),
        cgst_percent=to_decimal_or_none(row.get("CGST %")),
        sgst_percent=to_decimal_or_none(row.get("SGST %")),
        cess_percent=to_decimal_or_none(row.get("CESS %")),
        igst_amount=to_decimal(row.get("IGST Value")),
        cgst_amount=to_decimal(row.get("CGST Value")),
        sgst_amount=to_decimal(row.get("SGST Value")),
        cess_amount=to_decimal(row.get("CESS Value")),
        total_tax=to_decimal(row.get("Total Tax")),
        total_gross_bill_amount=to_decimal(row.get("Total Gross Bill Amount")),
        commission_percent=to_decimal_or_none(row.get("Commission %")),
        commission_charge=to_decimal(row.get("Commission Charge (Rs)")),
        commission_gst=to_decimal(row.get("Commission GST (Rs)")),
        shipping_charge=to_decimal(row.get("Shipping Charge (Rs)")),
        shipping_gst=to_decimal(row.get("Shipping GST (Rs)")),
        tcs_amount=to_decimal(row.get("TCS Amount")),
        tds_194o_amount=to_decimal(row.get("TDS 194O Amount")),
        tds_194q_amount=to_decimal(row.get("TDS 194Q Amount")),
        net_deductions=to_decimal(row.get("Net Deductions")),
        net_additions=to_decimal(row.get("Net Additions")),
        item_level_payout=to_decimal(row.get("Item Level Payout")),
        unsettled_amount=to_decimal(row.get("Unsettled Amount")),
        # COMPLETE ORIGINAL EXCEL ROW
        raw_data=serialize_raw_row(row),
        source_import=import_batch,
    )

    return {
        "order_created": order_created,
        "item_created": True,
        "skipped": False,
        "order_item": order_item,
    }


# ---------------------------------------------------------------------
# Return / Cancelled Orders
# ---------------------------------------------------------------------


def import_return_row(
    *,
    row,
    account,
    import_batch,
    payout,
):
    """
    Import one row from Cancelled or Returned Orders.

    In the current Blinkit report:
        - CANCELLED status -> order_kind = CANCELLED
        - other return rows -> order_kind = RETURN

    The return order ID is the same as the corresponding forward
    order ID in the uploaded report.
    """

    return_order_id = normalize_id(row.get("Return Order ID"))

    forward_invoice_id = normalize_text(row.get("Forward Invoice ID"))

    return_invoice_id = normalize_text(row.get("Return Invoice ID"))

    item_id = normalize_id(row.get("Item ID"))

    if not return_order_id:
        raise ValueError("Missing Return Order ID.")

    if not item_id:
        raise ValueError(f"Missing Item ID for Return Order {return_order_id}.")

    order_status = normalize_text(row.get("Order Status"))

    if order_status == "CANCELLED":
        order_kind = "CANCELLED"
    else:
        order_kind = "RETURN"

    return_order_date = parse_date(row.get("Return Order Date"))

    # The return sheet gives us the forward invoice.
    # Find the forward order so we can preserve the relationship
    # where possible.
    forward_order = BlinkitOrder.objects.filter(
        account=account,
        invoice_id=forward_invoice_id,
        order_kind="FORWARD",
    ).first()

    order, order_created = BlinkitOrder.objects.get_or_create(
        account=account,
        order_id=return_order_id,
        order_kind=order_kind,
        defaults={
            "invoice_id": return_invoice_id,
            "order_type": "return",
            "order_date": return_order_date,
            "order_status": order_status,
            "forward_invoice_id": forward_invoice_id,
            "return_invoice_id": return_invoice_id,
            "return_order_date": return_order_date,
            "customer_name": normalize_text(row.get("Customer Name")),
            "customer_gst_name": normalize_text(
                row.get("GST Name (where buyer has given gst)")
            ),
            "customer_gst_number": normalize_text(
                row.get("GST Number (where buyer has given gst)")
            ),
            "supply_state": normalize_text(row.get("Supply State")),
            "customer_city": normalize_text(row.get("Customer City")),
            "customer_state": normalize_text(row.get("Customer State")),
            "payout": payout,
            # COMPLETE ORIGINAL EXCEL ROW
            "raw_data": serialize_raw_row(row),
            "source_import": import_batch,
        },
    )

    if not order_created:
        changed = False

        if not order.forward_invoice_id and forward_invoice_id:
            order.forward_invoice_id = forward_invoice_id
            changed = True

        if not order.return_invoice_id and return_invoice_id:
            order.return_invoice_id = return_invoice_id
            changed = True

        if order.payout_id is None:
            order.payout = payout
            changed = True

        if order.source_import_id is None:
            order.source_import = import_batch
            changed = True

        if changed:
            order.save(
                update_fields=[
                    "forward_invoice_id",
                    "return_invoice_id",
                    "payout",
                    "source_import",
                    "updated_at",
                ]
            )

    # -------------------------------------------------------------
    # Duplicate return item protection
    # -------------------------------------------------------------

    existing_item = BlinkitOrderItem.objects.filter(
        order=order,
        item_id=item_id,
    ).first()

    if existing_item:
        return {
            "order_created": order_created,
            "item_created": False,
            "skipped": True,
        }

    product = find_product(
        account=account,
        item_id=item_id,
    )

    order_item = BlinkitOrderItem.objects.create(
        order=order,
        product=product,
        item_id=item_id,
        product_name=normalize_text(row.get("Product Name")),
        variant_description=normalize_text(row.get("Variant Description")),
        quantity=to_decimal(row.get("Quantity")),
        mrp=to_decimal_or_none(row.get("MRP (Rs)")),
        selling_price=to_decimal_or_none(row.get("Selling Price (Rs)")),
        igst_percent=to_decimal_or_none(row.get("IGST (%)")),
        cgst_percent=to_decimal_or_none(row.get("CGST (%)")),
        sgst_percent=to_decimal_or_none(row.get("SGST (%)")),
        cess_percent=to_decimal_or_none(row.get("Cess (%)")),
        igst_amount=to_decimal(row.get("IGST Value")),
        cgst_amount=to_decimal(row.get("CGST Value")),
        sgst_amount=to_decimal(row.get("SGST Value")),
        cess_amount=to_decimal(row.get("CESS Value")),
        total_tax=to_decimal(row.get("Total Tax")),
        total_gross_bill_amount=to_decimal(row.get("Total Gross Bill Amount")),
        commission_percent=to_decimal_or_none(row.get("Commission %")),
        commission_charge=to_decimal(row.get("Commission Charge (Rs)")),
        commission_gst=to_decimal(row.get("Commission GST (Rs)")),
        shipping_charge=to_decimal(row.get("Shipping Charge (Rs)")),
        shipping_gst=to_decimal(row.get("Shipping GST (Rs)")),
        tcs_amount=to_decimal(row.get("TCS Amount")),
        tds_194o_amount=to_decimal(row.get("TDS 194O Amount")),
        tds_194q_amount=to_decimal(row.get("TDS 194Q Amount")),
        net_deductions=to_decimal(row.get("Net Deductions")),
        # Return sheet does not contain Net Additions.
        net_additions=Decimal("0"),
        item_level_payout=to_decimal(row.get("Item Level Payout")),
        unsettled_amount=to_decimal(row.get("Unsettled Amount")),
        # COMPLETE ORIGINAL EXCEL ROW
        raw_data=serialize_raw_row(row),
        source_import=import_batch,
    )

    return {
        "order_created": order_created,
        "item_created": True,
        "skipped": False,
        "order_item": order_item,
    }


# ---------------------------------------------------------------------
# Main importer
# ---------------------------------------------------------------------


@transaction.atomic
def import_order_report(
    *,
    file,
    account,
    import_batch,
):
    """
    Import the complete Blinkit Order Financial workbook.

    Sheets:
        Forward Orders
        Cancelled or Returned Orders

    Payout Breakup is intentionally not imported.
    """

    file.seek(0)

    workbook = load_workbook(
        file,
        read_only=True,
        data_only=True,
    )

    if FORWARD_SHEET not in workbook.sheetnames:
        raise ValueError(f"Missing required sheet: {FORWARD_SHEET}")

    if RETURN_SHEET not in workbook.sheetnames:
        raise ValueError(f"Missing required sheet: {RETURN_SHEET}")

    forward_rows = list(
        get_sheet_rows(
            workbook[FORWARD_SHEET],
            header_row=5,
            data_start_row=7,
        )
    )

    return_rows = list(
        get_sheet_rows(
            workbook[RETURN_SHEET],
            header_row=5,
            data_start_row=7,
        )
    )

    if not forward_rows and not return_rows:
        raise ValueError("Blinkit order report contains no data.")

    # -------------------------------------------------------------
    # Determine payout information
    # -------------------------------------------------------------

    all_rows = forward_rows + return_rows

    payout_groups = {}

    for row in all_rows:
        settlement_date = parse_date(row.get("Settlement Date"))

        bank_utr = normalize_text(row.get("Bank UTR"))

        settlement_status = normalize_text(row.get("Settlement Status"))

        if not settlement_date:
            raise ValueError("Settlement Date is missing from the report.")

        group_key = (
            bank_utr,
            settlement_date.date(),
            settlement_status,
        )

        payout_groups[group_key] = {
            "settlement_date": settlement_date,
            "bank_utr": bank_utr,
            "settlement_status": settlement_status,
        }

    # -------------------------------------------------------------
    # Create payouts
    # -------------------------------------------------------------

    payouts = {}

    for key, payout_data in payout_groups.items():
        payout = get_or_create_payout(
            account=account,
            settlement_date=payout_data["settlement_date"],
            bank_utr=payout_data["bank_utr"],
            settlement_status=payout_data["settlement_status"],
        )

        payouts[key] = payout

    # -------------------------------------------------------------
    # Import forward orders
    # -------------------------------------------------------------

    forward_imported_orders = 0
    forward_imported_items = 0
    forward_skipped_items = 0

    for row in forward_rows:
        settlement_date = parse_date(row.get("Settlement Date"))

        group_key = (
            normalize_text(row.get("Bank UTR")),
            settlement_date.date(),
            normalize_text(row.get("Settlement Status")),
        )

        payout = payouts[group_key]

        result = import_forward_row(
            row=row,
            account=account,
            import_batch=import_batch,
            payout=payout,
        )

        if result["order_created"]:
            forward_imported_orders += 1

        if result["item_created"]:
            forward_imported_items += 1

        if result["skipped"]:
            forward_skipped_items += 1

    # -------------------------------------------------------------
    # Import returns/cancellations
    # -------------------------------------------------------------

    return_imported_orders = 0
    return_imported_items = 0
    return_skipped_items = 0

    for row in return_rows:
        settlement_date = parse_date(row.get("Settlement Date"))

        group_key = (
            normalize_text(row.get("Bank UTR")),
            settlement_date.date(),
            normalize_text(row.get("Settlement Status")),
        )

        payout = payouts[group_key]

        result = import_return_row(
            row=row,
            account=account,
            import_batch=import_batch,
            payout=payout,
        )

        if result["order_created"]:
            return_imported_orders += 1

        if result["item_created"]:
            return_imported_items += 1

        if result["skipped"]:
            return_skipped_items += 1

    # -------------------------------------------------------------
    # Update ImportBatch payout period
    # -------------------------------------------------------------

    settlement_dates = [
        data["settlement_date"].date() for data in payout_groups.values()
    ]

    if settlement_dates:
        # Current workbook contains one payout cycle.
        # If multiple cycles are uploaded in one file later,
        # this can be changed to separate batches.
        first_settlement = min(settlement_dates)

        period_start, period_end = derive_payout_period(
            datetime.combine(
                first_settlement,
                datetime.min.time(),
            )
        )

        import_batch.payout_period_start = period_start
        import_batch.payout_period_end = period_end

        import_batch.save(
            update_fields=[
                "payout_period_start",
                "payout_period_end",
            ]
        )

    return {
        "forward_rows": len(forward_rows),
        "return_rows": len(return_rows),
        "forward_orders_imported": forward_imported_orders,
        "forward_items_imported": forward_imported_items,
        "forward_items_skipped": forward_skipped_items,
        "return_orders_imported": return_imported_orders,
        "return_items_imported": return_imported_items,
        "return_items_skipped": return_skipped_items,
        "payouts_created_or_found": len(payouts),
    }
