# amazon_auth/listing_sync.py

from django.db import transaction

from amazon_auth.models import AmazonListingItem
from amazon_auth.spapi_manager import SPAPIManager


from django.db.models import Q

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from amazon_auth.serializers import AmazonListingItemSerializer

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font
from django.utils.timezone import localtime
import json

from openpyxl import load_workbook
from zipfile import BadZipFile

from openpyxl import load_workbook
from io import BytesIO
from openpyxl.utils import get_column_letter

import os
import uuid

from django.conf import settings
from django.http import JsonResponse
from django.utils.timezone import localtime

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter



def sync_listing_items(user, account):

    manager = SPAPIManager(user=user, account=account)

    seller_id = account.seller_central_id
    marketplace_id = account.marketplace_id

    next_token = None
    total_synced = 0

    while True:

        response = manager.search_listing_items(
            seller_id=seller_id,
            marketplace_id=marketplace_id,
            page_token=next_token
        )

        items = response.get("items", [])

        for item in items:

            try:

                sku = item.get("sku")

                summary = item.get("summaries", [{}])[0]

                asin = summary.get("asin")

                marketplace_id = summary.get("marketplaceId")

                product_type = summary.get("productType")

                condition_type = summary.get("conditionType")

                status = summary.get("status", [])

                fnsku = summary.get("fnSku")

                item_name = summary.get("itemName")

                created_date = summary.get("createdDate")

                last_updated_date = summary.get("lastUpdatedDate")

                main_image = summary.get("mainImage", {})

                image_url = main_image.get("link") if main_image else None

                with transaction.atomic():

                    AmazonListingItem.objects.update_or_create(

                        amazon_account=account,
                        sku=sku,
                        marketplace_id=marketplace_id,

                        defaults={

                            "user": user,
                            "asin": asin,
                            "product_type": product_type,
                            "condition_type": condition_type,
                            "status": status,
                            "fnsku": fnsku,
                            "item_name": item_name,
                            "image_url": image_url,
                            "created_date": created_date,
                            "last_updated_date": last_updated_date,
                            "attributes": item.get("attributes", {}),
                            "issues": item.get("issues", []),
                            "offers": item.get("offers", []),
                            "fulfillment_availability": item.get(
                                "fulfillmentAvailability",
                                []
                            ),
                            "relationships": item.get(
                                "relationships",
                                []
                            ),
                            "product_types": item.get(
                                "productTypes",
                                []
                            ),
                            "raw_response": item
                        }
                    )

                total_synced += 1

                print(f"Synced SKU: {sku}")

            except Exception as e:

                print(f"Error syncing SKU {item.get('sku')}: {str(e)}")

        pagination = response.get("pagination", {})

        next_token = pagination.get("nextToken")

        if not next_token:
            break

    return total_synced


class AmazonListingItemsView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            user = request.user

            data = request.data

            search = data.get("search", "")
            marketplace_id = data.get("marketplace_id")
            product_type = data.get("product_type")
            condition_type = data.get("condition_type")
            asin = data.get("asin")
            sku = data.get("sku")

            pagination = data.get("pagination", {})

            page_no = int(pagination.get("page", 1))
            page_size = int(pagination.get("page_size", 10))

            queryset = (
                AmazonListingItem.objects
                .filter(user=user)
                .select_related("amazon_account")
                .order_by("-last_updated_date")
            )

            # SEARCH
            if search:

                queryset = queryset.filter(
                    Q(sku__icontains=search) |
                    Q(asin__icontains=search) |
                    Q(item_name__icontains=search) |
                    Q(product_type__icontains=search) |
                    Q(fnsku__icontains=search)
                )

            # FILTERS
            if marketplace_id:

                queryset = queryset.filter(
                    marketplace_id=marketplace_id
                )

            if product_type:

                queryset = queryset.filter(
                    product_type__icontains=product_type
                )

            if condition_type:

                queryset = queryset.filter(
                    condition_type=condition_type
                )

            if asin:

                queryset = queryset.filter(
                    asin=asin
                )

            if sku:

                queryset = queryset.filter(
                    sku=sku
                )

            total_count = queryset.count()

            start = (page_no - 1) * page_size
            end = start + page_size

            queryset = queryset[start:end]

            serializer = AmazonListingItemSerializer(
                queryset,
                many=True
            )

            return Response({

                "status": True,
                "statusCode": 200,
                "message": "Amazon listing items fetched successfully",
                "totalCount": total_count,
                "pageNo": page_no,
                "pageSize": page_size,
                "totalPages": (
                    total_count + page_size - 1
                ) // page_size,

                "data": serializer.data

            }, status=status.HTTP_200_OK)

        except Exception as e:

            return Response({
                "status": False,
                "message": str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
        

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def full_export_amazon_listing_excel(request):

    user = request.user

    amazon_account_id = request.GET.get("amazon_account_id")
    sku = request.GET.get("sku")
    asin = request.GET.get("asin")
    marketplace_id = request.GET.get("marketplace_id")

    queryset = AmazonListingItem.objects.filter(user=user)

    # ---------------- FILTERS ----------------

    if amazon_account_id:
        queryset = queryset.filter(amazon_account_id=amazon_account_id)

    if sku:
        queryset = queryset.filter(sku__icontains=sku)

    if asin:
        queryset = queryset.filter(asin__icontains=asin)

    if marketplace_id:
        queryset = queryset.filter(marketplace_id=marketplace_id)

    queryset = queryset.order_by("-id")

    # ---------------- CREATE EXCEL ----------------

    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Listing Items"

    # ---------------- HEADERS ----------------

    headers = [
        "SKU",
        "ASIN",
        "Marketplace ID",
        "Product Type",
        "Condition Type",
        "FNSKU",
        "Item Name",
        "Image URL",

        # NEW FIELDS
        "Standard Cost",
        "GST Rate",
        "TCS",
        "Region",
        "Shipping Estimate",
        "Step Level",

        "Created Date",
        "Last Updated Date",
        "Status",
        "Issues",
        "Offers",
        "Fulfillment Availability",
        "Relationships",
        "Product Types",
        "Attributes",
        "Created At",
        "Updated At",
    ]

    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # ---------------- DATA ----------------

    for row_num, item in enumerate(queryset, start=2):

        ws.cell(row=row_num, column=1, value=item.sku)
        ws.cell(row=row_num, column=2, value=item.asin)
        ws.cell(row=row_num, column=3, value=item.marketplace_id)
        ws.cell(row=row_num, column=4, value=item.product_type)
        ws.cell(row=row_num, column=5, value=item.condition_type)
        ws.cell(row=row_num, column=6, value=item.fnsku)
        ws.cell(row=row_num, column=7, value=item.item_name)
        ws.cell(row=row_num, column=8, value=item.image_url)

        # NEW FIELDS
        ws.cell(row=row_num, column=9, value=item.standard_cost)
        ws.cell(row=row_num, column=10, value=item.gst_rate)
        ws.cell(row=row_num, column=11, value=item.tcs)
        ws.cell(row=row_num, column=12, value=item.region)
        ws.cell(row=row_num, column=13, value=item.shiping_estimate)
        ws.cell(row=row_num, column=14, value=item.step_level)

        ws.cell(
            row=row_num,
            column=15,
            value=localtime(item.created_date).strftime("%Y-%m-%d %H:%M:%S")
            if item.created_date else ""
        )

        ws.cell(
            row=row_num,
            column=16,
            value=localtime(item.last_updated_date).strftime("%Y-%m-%d %H:%M:%S")
            if item.last_updated_date else ""
        )

        ws.cell(row=row_num, column=17, value=json.dumps(item.status))
        ws.cell(row=row_num, column=18, value=json.dumps(item.issues))
        ws.cell(row=row_num, column=19, value=json.dumps(item.offers))
        ws.cell(row=row_num, column=20, value=json.dumps(item.fulfillment_availability))
        ws.cell(row=row_num, column=21, value=json.dumps(item.relationships))
        ws.cell(row=row_num, column=22, value=json.dumps(item.product_types))
        ws.cell(row=row_num, column=23, value=json.dumps(item.attributes))

        ws.cell(
            row=row_num,
            column=24,
            value=localtime(item.created_at).strftime("%Y-%m-%d %H:%M:%S")
        )

        ws.cell(
            row=row_num,
            column=25,
            value=localtime(item.updated_at).strftime("%Y-%m-%d %H:%M:%S")
        )

    # ---------------- AUTO WIDTH ----------------

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 5, 50)

    # ---------------- RESPONSE ----------------

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="amazon_listing_items.xlsx"'

    wb.save(response)

    return response


# @api_view(["GET"])
# @permission_classes([IsAuthenticated])
# def export_amazon_listing_excel(request):

#     user = request.user

#     amazon_account_id = request.GET.get("amazon_account_id")
#     sku = request.GET.get("sku")
#     asin = request.GET.get("asin")
#     marketplace_id = request.GET.get("marketplace_id")

#     queryset = AmazonListingItem.objects.filter(user=user)

#     # ---------------- FILTERS ----------------

#     if amazon_account_id:
#         queryset = queryset.filter(amazon_account_id=amazon_account_id)

#     if sku:
#         queryset = queryset.filter(sku__icontains=sku)

#     if asin:
#         queryset = queryset.filter(asin__icontains=asin)

#     if marketplace_id:
#         queryset = queryset.filter(marketplace_id=marketplace_id)

#     queryset = queryset.order_by("-last_updated_date")

#     # ---------------- CREATE EXCEL ----------------

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Amazon Listing Items"

#     # ---------------- HEADERS ----------------

#     headers = [
#         "SKU",
#         "ASIN",

#         # WEIGHT
#         "Item Weight",
#         "Item Weight Unit",

#         # PACKAGE WEIGHT
#         "Item Pkg Weight",
#         "Package Weight Unit",

#         # ITEM DIMENSIONS
#         "Length",
#         "Width",
#         "Height",
#         "Dimension Unit",

#         # PACKAGE DIMENSIONS
#         "Pkg Length",
#         "Pkg Width",
#         "Pkg Height",
#         "Pkg Dimension Unit",

#         "Shipping Estimate Charges",
#         "Region",
#         "Step Level",

#         # COST FIELDS
#         "Product Cost",
#         "GST Rate",
#         "TCS",

#         # "Created At",
#         # "Updated At",
#     ]

#     # ---------------- ADD HEADERS ----------------

#     for col_num, header in enumerate(headers, 1):
#         cell = ws.cell(row=1, column=col_num)
#         cell.value = header
#         cell.font = Font(bold=True)

#     # ---------------- DATA ----------------

#     for row_num, item in enumerate(queryset, start=2):

#         attributes = item.attributes or {}

#         # ---------------- ITEM DIMENSIONS ----------------

#         item_dimensions = attributes.get("item_dimensions", [])

#         item_length = ""
#         item_width = ""
#         item_height = ""
#         item_dimension_unit = ""

#         if item_dimensions and isinstance(item_dimensions, list):

#             dimension_data = item_dimensions[0]

#             item_length = dimension_data.get("length", {}).get("value", "")
#             item_width = dimension_data.get("width", {}).get("value", "")
#             item_height = dimension_data.get("height", {}).get("value", "")
#             item_dimension_unit = dimension_data.get("length", {}).get("unit", "")

#         # ---------------- PACKAGE DIMENSIONS ----------------

#         package_dimensions = attributes.get("item_package_dimensions", [])

#         pkg_length = ""
#         pkg_width = ""
#         pkg_height = ""
#         pkg_dimension_unit = ""

#         if package_dimensions and isinstance(package_dimensions, list):

#             pkg_data = package_dimensions[0]

#             pkg_length = pkg_data.get("length", {}).get("value", "")
#             pkg_width = pkg_data.get("width", {}).get("value", "")
#             pkg_height = pkg_data.get("height", {}).get("value", "")
#             pkg_dimension_unit = pkg_data.get("length", {}).get("unit", "")

#         # ---------------- ITEM WEIGHT ----------------

#         item_weight = attributes.get("item_weight", [])

#         weight_value = ""
#         weight_unit = ""

#         if item_weight and isinstance(item_weight, list):

#             weight_data = item_weight[0]

#             weight_value = weight_data.get("value", "")
#             weight_unit = weight_data.get("unit", "")

#         # ---------------- PACKAGE WEIGHT ----------------

#         item_package_weight = attributes.get("item_package_weight", [])

#         package_weight_value = ""
#         package_weight_unit = ""

#         if item_package_weight and isinstance(item_package_weight, list):

#             package_data = item_package_weight[0]

#             package_weight_value = package_data.get("value", "")
#             package_weight_unit = package_data.get("unit", "")

#         # ---------------- WRITE DATA ----------------

#         ws.cell(row=row_num, column=1, value=item.sku)
#         ws.cell(row=row_num, column=2, value=item.asin)

#         # ITEM WEIGHT
#         ws.cell(row=row_num, column=3, value=weight_value)
#         ws.cell(row=row_num, column=4, value=weight_unit)

#         # PACKAGE WEIGHT
#         ws.cell(row=row_num, column=5, value=package_weight_value)
#         ws.cell(row=row_num, column=6, value=package_weight_unit)

#         # ITEM DIMENSIONS
#         ws.cell(row=row_num, column=7, value=item_length)
#         ws.cell(row=row_num, column=8, value=item_width)
#         ws.cell(row=row_num, column=9, value=item_height)
#         ws.cell(row=row_num, column=10, value=item_dimension_unit)

#         # PACKAGE DIMENSIONS
#         ws.cell(row=row_num, column=11, value=pkg_length)
#         ws.cell(row=row_num, column=12, value=pkg_width)
#         ws.cell(row=row_num, column=13, value=pkg_height)
#         ws.cell(row=row_num, column=14, value=pkg_dimension_unit)

#         # OTHER FIELDS
#         ws.cell(row=row_num, column=15, value=item.shiping_estimate)
#         ws.cell(row=row_num, column=16, value=item.region)
#         ws.cell(row=row_num, column=17, value=item.step_level)

#         # COST FIELDS
#         ws.cell(row=row_num, column=18, value=item.standard_cost)
#         ws.cell(row=row_num, column=19, value=item.gst_rate)
#         ws.cell(row=row_num, column=20, value=item.tcs)

#         # DATES
#         # ws.cell(
#         #     row=row_num,
#         #     column=21,
#         #     value=localtime(item.created_at).strftime("%Y-%m-%d %H:%M:%S")
#         #     if item.created_at else ""
#         # )

#         # ws.cell(
#         #     row=row_num,
#         #     column=22,
#         #     value=localtime(item.updated_at).strftime("%Y-%m-%d %H:%M:%S")
#         #     if item.updated_at else ""
#         # )

#     # ---------------- AUTO WIDTH ----------------

#     for column_cells in ws.columns:

#         max_length = max(
#             len(str(cell.value)) if cell.value else 0
#             for cell in column_cells
#         )

#         adjusted_width = min(max_length + 5, 50)

#         ws.column_dimensions[
#             column_cells[0].column_letter
#         ].width = adjusted_width

#     # ---------------- RESPONSE ----------------

#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )

#     response[
#         "Content-Disposition"
#     ] = 'attachment; filename="amazon_listing_items.xlsx"'

#     wb.save(response)

#     return response



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_amazon_listing_excel(request):

    user = request.user

    amazon_account_id = request.GET.get("amazon_account_id")
    sku = request.GET.get("sku")
    asin = request.GET.get("asin")
    marketplace_id = request.GET.get("marketplace_id")

    queryset = AmazonListingItem.objects.filter(
        user=user
    )

    # ============================================================
    # FILTERS
    # ============================================================

    if amazon_account_id:
        queryset = queryset.filter(
            amazon_account_id=amazon_account_id
        )

    if sku:
        queryset = queryset.filter(
            sku__icontains=sku
        )

    if asin:
        queryset = queryset.filter(
            asin__icontains=asin
        )

    if marketplace_id:
        queryset = queryset.filter(
            marketplace_id=marketplace_id
        )

    queryset = queryset.order_by(
        "-last_updated_date"
    )

    # ============================================================
    # CREATE WORKBOOK
    # ============================================================

    wb = Workbook()

    ws = wb.active
    ws.title = "Amazon Listing Items"

    # ============================================================
    # HEADERS
    # ============================================================

    headers = [

        "SKU",
        "ASIN",

        "Item Weight",
        "Item Weight Unit",

        "Package Weight",
        "Package Weight Unit",

        "Item Length",
        "Item Width",
        "Item Height",
        "Item Dimension Unit",

        "Package Length",
        "Package Width",
        "Package Height",
        "Package Dimension Unit",

        "Shipping Estimate",
        "Region",
        "Step Level",

        "Product Cost",
        "GST Rate",
        "TCS",
    ]

    # ============================================================
    # WRITE HEADERS
    # ============================================================

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(
            row=1,
            column=col_num,
            value=header
        )

        cell.font = Font(bold=True)

    # ============================================================
    # WRITE DATA
    # ============================================================

    for row_num, item in enumerate(queryset, start=2):

        try:

            attributes = item.attributes or {}

            # ----------------------------------------------------
            # ITEM DIMENSIONS
            # ----------------------------------------------------

            item_dimensions = attributes.get(
                "item_dimensions",
                []
            )

            item_length = ""
            item_width = ""
            item_height = ""
            item_dimension_unit = ""

            if item_dimensions and isinstance(item_dimensions, list):

                dimension_data = item_dimensions[0] or {}

                item_length = (
                    dimension_data.get("length", {})
                    .get("value", "")
                )

                item_width = (
                    dimension_data.get("width", {})
                    .get("value", "")
                )

                item_height = (
                    dimension_data.get("height", {})
                    .get("value", "")
                )

                item_dimension_unit = (
                    dimension_data.get("length", {})
                    .get("unit", "")
                )

            # ----------------------------------------------------
            # PACKAGE DIMENSIONS
            # ----------------------------------------------------

            package_dimensions = attributes.get(
                "item_package_dimensions",
                []
            )

            pkg_length = ""
            pkg_width = ""
            pkg_height = ""
            pkg_dimension_unit = ""

            if package_dimensions and isinstance(package_dimensions, list):

                pkg_data = package_dimensions[0] or {}

                pkg_length = (
                    pkg_data.get("length", {})
                    .get("value", "")
                )

                pkg_width = (
                    pkg_data.get("width", {})
                    .get("value", "")
                )

                pkg_height = (
                    pkg_data.get("height", {})
                    .get("value", "")
                )

                pkg_dimension_unit = (
                    pkg_data.get("length", {})
                    .get("unit", "")
                )

            # ----------------------------------------------------
            # ITEM WEIGHT
            # ----------------------------------------------------

            item_weight = attributes.get(
                "item_weight",
                []
            )

            weight_value = ""
            weight_unit = ""

            if item_weight and isinstance(item_weight, list):

                weight_data = item_weight[0] or {}

                weight_value = weight_data.get(
                    "value",
                    ""
                )

                weight_unit = weight_data.get(
                    "unit",
                    ""
                )

            # ----------------------------------------------------
            # PACKAGE WEIGHT
            # ----------------------------------------------------

            item_package_weight = attributes.get(
                "item_package_weight",
                []
            )

            package_weight_value = ""
            package_weight_unit = ""

            if item_package_weight and isinstance(item_package_weight, list):

                package_data = item_package_weight[0] or {}

                package_weight_value = package_data.get(
                    "value",
                    ""
                )

                package_weight_unit = package_data.get(
                    "unit",
                    ""
                )

            # ----------------------------------------------------
            # ROW DATA
            # ----------------------------------------------------

            row_data = [

                str(item.sku or ""),
                str(item.asin or ""),

                weight_value,
                weight_unit,

                package_weight_value,
                package_weight_unit,

                item_length,
                item_width,
                item_height,
                item_dimension_unit,

                pkg_length,
                pkg_width,
                pkg_height,
                pkg_dimension_unit,

                float(item.shiping_estimate or 0),
                str(item.region or ""),
                str(item.step_level or ""),

                float(item.standard_cost or 0),
                float(item.gst_rate or 0),
                float(item.tcs or 0),
            ]

            for col_num, value in enumerate(row_data, 1):

                ws.cell(
                    row=row_num,
                    column=col_num,
                    value=value
                )

        except Exception as e:

            print(
                f"Excel export error for SKU "
                f"{item.sku}: {e}"
            )

    # ============================================================
    # AUTO WIDTH
    # ============================================================

    for column_cells in ws.columns:

        max_length = 0

        column_letter = get_column_letter(
            column_cells[0].column
        )

        for cell in column_cells:

            try:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except:
                pass

        adjusted_width = min(
            max_length + 5,
            50
        )

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    # ============================================================
    # CREATE EXPORT FOLDER
    # ============================================================

    export_dir = os.path.join(
        settings.MEDIA_ROOT,
        "exports"
    )

    os.makedirs(
        export_dir,
        exist_ok=True
    )

    # ============================================================
    # FILE NAME
    # ============================================================

    file_name = (
        f"amazon_listing_"
        f"{uuid.uuid4().hex}.xlsx"
    )

    file_path = os.path.join(
        export_dir,
        file_name
    )

    # ============================================================
    # SAVE FILE
    # ============================================================

    wb.save(file_path)

    # ============================================================
    # FILE URL
    # ============================================================

    file_url = (
        request.build_absolute_uri(
            settings.MEDIA_URL
        )
        + f"exports/{file_name}"
    )

    # ============================================================
    # RESPONSE
    # ============================================================

    return JsonResponse({
        "status": True,
        "message": "Excel exported successfully",
        "download_url": file_url
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def upload_amazon_listing_excel(request):

    user = request.user

    excel_file = request.FILES.get("file")

    if not excel_file:
        return Response(
            {
                "success": False,
                "message": "Excel file is required"
            },
            status=400
        )

    # =========================================================
    # FILE VALIDATION
    # =========================================================

    if not excel_file.name.endswith(".xlsx"):

        return Response(
            {
                "success": False,
                "message": "Only .xlsx files are supported"
            },
            status=400
        )

    try:

        workbook = load_workbook(excel_file)

    except BadZipFile:

        return Response(
            {
                "success": False,
                "message": "Invalid or corrupted Excel file"
            },
            status=400
        )

    except Exception as e:

        return Response(
            {
                "success": False,
                "message": f"Excel read error: {str(e)}"
            },
            status=400
        )

    try:

        sheet = workbook.active

        # =========================================================
        # HEADER MAP
        # =========================================================

        headers = {}

        for col in range(1, sheet.max_column + 1):

            header_value = sheet.cell(row=1, column=col).value

            if header_value:
                headers[str(header_value).strip()] = col

        # =========================================================
        # REQUIRED HEADERS
        # =========================================================

        required_headers = [
            "SKU",
            "ASIN"
        ]

        missing_headers = [
            header
            for header in required_headers
            if header not in headers
        ]

        if missing_headers:

            return Response(
                {
                    "success": False,
                    "message": f"Missing headers: {', '.join(missing_headers)}"
                },
                status=400
            )

        updated_count = 0
        skipped_count = 0
        skipped_rows = []

        # =========================================================
        # HELPER FUNCTION
        # =========================================================

        def get_cell(row, header_name, default=None):

            column = headers.get(header_name)

            if not column:
                return default

            value = sheet.cell(row=row, column=column).value

            if value is None:
                return default

            return value

        # =========================================================
        # PROCESS ROWS
        # =========================================================

        with transaction.atomic():

            for row in range(2, sheet.max_row + 1):

                try:

                    sku = get_cell(row, "SKU")
                    asin = get_cell(row, "ASIN")

                    if not sku or not asin:

                        skipped_count += 1

                        skipped_rows.append({
                            "row": row,
                            "reason": "SKU or ASIN missing"
                        })

                        continue

                    sku = str(sku).strip()
                    asin = str(asin).strip()

                    # =================================================
                    # FIND ITEM
                    # =================================================

                    item = AmazonListingItem.objects.filter(
                        user=user,
                        sku=sku,
                        asin=asin
                    ).first()

                    if not item:

                        skipped_count += 1

                        skipped_rows.append({
                            "row": row,
                            "sku": sku,
                            "asin": asin,
                            "reason": "Listing item not found"
                        })

                        continue

                    attributes = item.attributes or {}

                    # =================================================
                    # ITEM WEIGHT
                    # =================================================

                    item_weight_value = get_cell(row, "Item Weight")
                    item_weight_unit = get_cell(row, "Item Weight Unit", "")

                    if (
                        item_weight_value is not None
                        or item_weight_unit
                    ):

                        attributes["item_weight"] = [
                            {
                                "value": float(item_weight_value or 0),
                                "unit": str(item_weight_unit or ""),
                                "marketplace_id": item.marketplace_id
                            }
                        ]

                    # =================================================
                    # PACKAGE WEIGHT
                    # =================================================

                    package_weight_value = get_cell(row, "Item Pkg Weight")
                    package_weight_unit = get_cell(row, "Package Weight Unit", "")

                    if (
                        package_weight_value is not None
                        or package_weight_unit
                    ):

                        attributes["item_package_weight"] = [
                            {
                                "value": float(package_weight_value or 0),
                                "unit": str(package_weight_unit or ""),
                                "marketplace_id": item.marketplace_id
                            }
                        ]

                    # =================================================
                    # ITEM DIMENSIONS
                    # =================================================

                    length = get_cell(row, "Length")
                    width = get_cell(row, "Width")
                    height = get_cell(row, "Height")
                    dimension_unit = get_cell(row, "Dimension Unit", "")

                    if (
                        length is not None
                        or width is not None
                        or height is not None
                    ):

                        attributes["item_dimensions"] = [
                            {
                                "length": {
                                    "value": float(length or 0),
                                    "unit": str(dimension_unit)
                                },
                                "width": {
                                    "value": float(width or 0),
                                    "unit": str(dimension_unit)
                                },
                                "height": {
                                    "value": float(height or 0),
                                    "unit": str(dimension_unit)
                                },
                                "marketplace_id": item.marketplace_id
                            }
                        ]

                    # =================================================
                    # PACKAGE DIMENSIONS
                    # =================================================

                    pkg_length = get_cell(row, "Pkg Length")
                    pkg_width = get_cell(row, "Pkg Width")
                    pkg_height = get_cell(row, "Pkg Height")
                    pkg_dimension_unit = get_cell(
                        row,
                        "Pkg Dimension Unit",
                        ""
                    )

                    if (
                        pkg_length is not None
                        or pkg_width is not None
                        or pkg_height is not None
                    ):

                        attributes["item_package_dimensions"] = [
                            {
                                "length": {
                                    "value": float(pkg_length or 0),
                                    "unit": str(pkg_dimension_unit)
                                },
                                "width": {
                                    "value": float(pkg_width or 0),
                                    "unit": str(pkg_dimension_unit)
                                },
                                "height": {
                                    "value": float(pkg_height or 0),
                                    "unit": str(pkg_dimension_unit)
                                },
                                "marketplace_id": item.marketplace_id
                            }
                        ]

                    # =================================================
                    # OTHER FIELDS
                    # =================================================

                    region = get_cell(row, "Region")
                    step_level = get_cell(row, "Step Level")

                    product_cost = get_cell(row, "Product Cost")
                    gst_rate = get_cell(row, "GST Rate")
                    tcs = get_cell(row, "TCS")

                    # =================================================
                    # UPDATE MODEL
                    # =================================================

                    if region is not None:
                        item.region = str(region)

                    if step_level is not None:
                        item.step_level = str(step_level)

                    if product_cost is not None:
                        item.standard_cost = float(product_cost or 0)

                    if gst_rate is not None:
                        item.gst_rate = float(gst_rate or 0)

                    if tcs is not None:
                        item.tcs = float(tcs or 0)

                    item.attributes = attributes

                    item.save()

                    updated_count += 1

                except Exception as row_error:

                    skipped_count += 1

                    skipped_rows.append({
                        "row": row,
                        "sku": str(sku) if sku else "",
                        "asin": str(asin) if asin else "",
                        "reason": str(row_error)
                    })

        # =========================================================
        # RESPONSE
        # =========================================================

        return Response(
            {
                "success": True,
                "message": "Excel uploaded successfully",
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "skipped_rows": skipped_rows
            }
        )

    except Exception as e:

        return Response(
            {
                "success": False,
                "message": str(e)
            },
            status=400
        )
    
