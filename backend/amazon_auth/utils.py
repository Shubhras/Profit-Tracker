import json
import time

# amazon_auth/date_utils.py (or wherever shared helpers live)
from datetime import datetime, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

import pandas as pd
from django.apps import apps
from openpyxl import Workbook

from .models import *

IST = ZoneInfo("Asia/Kolkata")
UTC = ZoneInfo("UTC")

#  IMPORTANT: change app name here
from amazon_auth.models import Order


# ---------------------------
# HELPER: RAW DATA PARSER
# ---------------------------
def extract_financials(raw_data):
    result = {
        "revenue": Decimal('0.00'),
        "fees": Decimal('0.00'),
        "tds": Decimal('0.00'),
        "promotions": Decimal('0.00'),
        "other": Decimal('0.00')
    }

    try:
        data = json.loads(raw_data) if isinstance(raw_data, str) else raw_data

        for item in data.get("ShipmentItemList", []):

            # Revenue
            for charge in item.get("ItemChargeList", []):
                amt = Decimal(str(charge.get("ChargeAmount", {}).get("CurrencyAmount", 0)))
                if charge.get("ChargeType") in ["Principal", "Tax"]:
                    result["revenue"] += amt
                else:
                    result["other"] += amt

            # Fees
            for fee in item.get("ItemFeeList", []):
                amt = Decimal(str(fee.get("FeeAmount", {}).get("CurrencyAmount", 0)))
                result["fees"] += abs(amt)

            # TDS
            for tax in item.get("ItemTaxWithheldList", []):
                for t in tax.get("TaxesWithheld", []):
                    amt = Decimal(str(t.get("ChargeAmount", {}).get("CurrencyAmount", 0)))
                    result["tds"] += abs(amt)

            # Promotions
            for promo in item.get("PromotionList", []):
                amt = Decimal(str(promo.get("PromotionAmount", {}).get("CurrencyAmount", 0)))
                result["promotions"] += abs(amt)

    except Exception:
        pass

    return result

# correct one 
# def _get_sku_profits_for_dashboard(user, start_date, end_date, filters={}):
def _get_sku_profits_for_dashboard(user, start_date, end_date, filters={}, from_date_ist=None, to_date_ist=None):
    from amazon_auth.models import ProfitCalculationSetting
    try:
        profit_setting, _ = ProfitCalculationSetting.objects.get_or_create(user=user)
    except Exception:
        profit_setting = ProfitCalculationSetting.objects.filter(user=user).first()
        if not profit_setting:
            profit_setting = ProfitCalculationSetting(user=user)
    
    from django.db.models import (
        Avg,
        Case,
        DecimalField,
        F,
        Max,
        OuterRef,
        Q,
        Subquery,
        Sum,
        When,
    )

    from amazon_ads.models import ProductAdMetric
    from amazon_auth.models import (
        AmazonEstimatedFee,
        AmazonListingItem,
        AmazonTransaction,
        AmazonTransactionBreakdown,
        AmazonTransactionRelatedIdentifier,
        OrderItem,
    )
    
    order_filter = Q(order__user=user)
    if start_date:
        order_filter &= Q(order__purchase_date__gte=start_date)
    if end_date:
        order_filter &= Q(order__purchase_date__lte=end_date)
        
    channels = filters.get("channel", {}).get("IN", [])
    if channels:
        CHANNEL_MAP = {"Amazon-India": "A21TJRUUN4KGV"}
        marketplace_ids = [CHANNEL_MAP[ch] for ch in channels if ch in CHANNEL_MAP]
        if marketplace_ids:
            order_filter &= Q(order__marketplace_id__in=marketplace_ids)
            
    parent_ids = filters.get("parentproductid", {}).get("IN", []) if isinstance(filters.get("parentproductid"), dict) else []
    if parent_ids:
        order_filter &= Q(asin__in=parent_ids)

    search_term = filters.get("search") or filters.get("searchTerm") or filters.get("q") or filters.get("sku")
    if isinstance(search_term, list) and search_term:
        search_term = search_term[0]
    if search_term:
        search_term = str(search_term).strip()
        order_filter &= (
            Q(seller_sku__icontains=search_term) |
            Q(asin__icontains=search_term) |
            Q(title__icontains=search_term)
        )
        
    from_date = start_date
    to_date = end_date

    sku_results = []
# ---------------- ORDER ITEM AGG ----------------

    listing_qs = AmazonListingItem.objects.filter(
            user=user
        ).filter(
            Q(asin=OuterRef("parent_asin")) | Q(asin=OuterRef("asin")) | Q(sku=OuterRef("seller_sku"))
        ).order_by("-updated_at")
    
    items = (
        OrderItem.objects
        .filter(order_filter)
        .exclude(order__order_status__icontains='Cancel')

        .annotate(

            sku_standard_cost=Subquery(
                listing_qs.values("standard_cost")[:1]
            ),

            sku_gst_rate=Subquery(
                listing_qs.values("gst_rate")[:1]
            ),

            sku_tcs_rate=Subquery(
                listing_qs.values("tcs")[:1]
            ),

            sku_tds_rate=Subquery(
                listing_qs.values("tds")[:1]
            ),

            sku_region=Subquery(
                listing_qs.values("region")[:1]
            ),
        )

        .values('parent_asin')

        .annotate(
            title=Max('title'),
            image_url=Max('image_url'),

            grossqty=Sum('quantity_ordered'),
            quantity_shipped=Sum('quantity_shipped'),

            shipping_income=Sum('shipping_income'),
            shipping_price=Sum('shipping_price'),

            discount=Sum('discount'),
            promotion_discount=Sum('promotion_discount'),

            # avg_cost=Avg('item_price'),
            avg_cost=Avg(
                Case(
                    When(Q(order__order_status__icontains='Pending') & Q(item_price=0), then=F('new_item_price')),
                    default=F('item_price'),
                    output_field=DecimalField(max_digits=12, decimal_places=2)
                )
            ),

            item_tax=Sum('item_tax'),

            # grosssales=Sum('item_price'),
            grosssales=Sum(
                Case(
                    When(Q(order__order_status__icontains='Pending') & Q(item_price=0), then=F('new_item_price')),
                    default=F('item_price'),
                    output_field=DecimalField(max_digits=12, decimal_places=2)
                )
            ),

            sku_standard_cost=Max('sku_standard_cost'),
            sku_gst_rate=Max('sku_gst_rate'),
            sku_tcs_rate=Max('sku_tcs_rate'),
            sku_tds_rate=Max('sku_tds_rate'),
            sku_region=Max('sku_region'),
        )
    )

    # ---------------- ESTIMATED FEES ----------------
    estimated_fee_qs = AmazonEstimatedFee.objects.filter(
        order_item__order__user=user
    ).exclude(order_item__order__order_status__icontains='Cancel')

    # apply same date filter
    if from_date:
        estimated_fee_qs = estimated_fee_qs.filter(
            order_item__order__purchase_date__gte=from_date
        )

    if to_date:
        estimated_fee_qs = estimated_fee_qs.filter(
            order_item__order__purchase_date__lte=to_date
        )

    # apply same parent filter
    if parent_ids:
        estimated_fee_qs = estimated_fee_qs.filter(
            order_item__parent_asin__in=parent_ids
        )

    estimated_fee_data = (
        estimated_fee_qs
        .values('asin')
        .annotate(
            estimated_fees=Sum('total_fees'),

            referral_fee=Sum('referral_fee'),
            closing_fee=Sum('closing_fee'),
            per_item_fee=Sum('per_item_fee'),

            fba_fee=Sum('fba_fee'),
            fba_pick_pack_fee=Sum('fba_pick_pack_fee'),
            fba_weight_handling_fee=Sum('fba_weight_handling_fee'),

            tax_amount=Sum('tax_amount'),
        )
    )

    estimated_fee_by_asin = {
        row['asin']: {
            "estimated_fees": float(row['estimated_fees'] or 0),
            "referral_fee": float(row['referral_fee'] or 0),
            "closing_fee": float(row['closing_fee'] or 0),
            "per_item_fee": float(row['per_item_fee'] or 0),
            "fba_fee": float(row['fba_fee'] or 0),
            "fba_pick_pack_fee": float(row['fba_pick_pack_fee'] or 0),
            "fba_weight_handling_fee": float(row['fba_weight_handling_fee'] or 0),
            "tax_amount": float(row['tax_amount'] or 0),
        }
        for row in estimated_fee_data
    }

    # ---------------- FINANCIAL EVENTS ----------------
    finances_qs = FinancialEvent.objects.filter(user=user)

    raw_map = (
        FinancialEvent.objects
        .filter(user=user)
        .exclude(raw_data=None)
        .values('amazon_order_id', 'raw_data')
    )

    raw_data_map = {}
    for r in raw_map:
        raw_data_map.setdefault(r['amazon_order_id'], []).append(r['raw_data'])

    if from_date:
        finances_qs = finances_qs.filter(posted_date__gte=from_date)
    if to_date:
        finances_qs = finances_qs.filter(posted_date__lte=to_date)

    finance_data = (
        finances_qs
        .values('amazon_order_id')
        .annotate(
            refund=Sum('total_amount', filter=Q(event_group="REFUND")),
            rto=Sum('total_amount', filter=Q(event_group="RTO")),
            # ads=Sum('total_amount', filter=Q(event_type__icontains='Ad')),
            commission=Sum('commission_fee'),
            fulfillment=Sum('fulfillment_fee'),
            other_fee=Sum('other_fee'),
            shipping_fee=Sum('shipping_fee'),
            gst=Sum('tax'),
        )
    )

    finance_map = {f['amazon_order_id']: f for f in finance_data}

    # ---------------- ASIN → ORDER MAP ----------------
    asin_orders = (
        OrderItem.objects
        .filter(order_filter)
        .exclude(order__order_status__icontains='Cancel')
        .values('asin','parent_asin', 'order__amazon_order_id', 'quantity_ordered', 'item_price','new_item_price','item_tax', 'promotion_discount')
    )

    child_parent_map = {}
    asin_map = {}
    for row in asin_orders:
        p_asin = row['parent_asin'] or row['asin']
        child_parent_map[row['asin']] = p_asin
        asin_map.setdefault(p_asin, []).append(row)

    estimated_fee_map = {}
    for asin, fee_data in estimated_fee_by_asin.items():
        p_asin = child_parent_map.get(asin) or asin
        if p_asin not in estimated_fee_map:
            estimated_fee_map[p_asin] = {
                "estimated_fees": 0.0, "referral_fee": 0.0, "closing_fee": 0.0,
                "per_item_fee": 0.0, "fba_fee": 0.0, "fba_pick_pack_fee": 0.0,
                "fba_weight_handling_fee": 0.0, "tax_amount": 0.0
            }
        for k, v in fee_data.items():
            estimated_fee_map[p_asin][k] += v

    # ---------------- TRANSACTION SHIPPING FEES ----------------

    matching_order_ids = [row["order__amazon_order_id"] for row in asin_orders]

    tx_identifiers = AmazonTransactionRelatedIdentifier.objects.filter(
        identifier_name="ORDER_ID", identifier_value__in=matching_order_ids
    ).values("transaction_id", "identifier_value")

    tx_to_order = {
        row["transaction_id"]: row["identifier_value"] for row in tx_identifiers
    }

    # ------------------------------------------------------------
    # SHIPPING STATUS PRIORITY
    # ------------------------------------------------------------
    #
    # Same financial event can appear as:
    #
    # DEFERRED
    # DEFERRED_RELEASED
    # RELEASED
    #
    # We use only the highest-priority lifecycle state.
    #
    # DEFERRED > DEFERRED_RELEASED > RELEASED
    # ------------------------------------------------------------

    STATUS_PRIORITY = {
        "DEFERRED": 3,
        "DEFERRED_RELEASED": 2,
        "RELEASED": 1,
    }

    tx_shipping_candidates = {}

    # ------------------------------------------------------------
    # MFN SHIPPING
    # ------------------------------------------------------------

    mfn_postage_txns = AmazonTransaction.objects.filter(
        id__in=tx_to_order.keys(),
        transaction_type="ServiceFee",
        transaction_status__in=[
            "DEFERRED",
            "DEFERRED_RELEASED",
            "RELEASED",
        ],
        description__icontains="MfnPostageFee",
    ).values(
        "id",
        "total_amount",
        "transaction_status",
    )

    for txn in mfn_postage_txns:
        order_id = tx_to_order.get(txn["id"])

        if not order_id:
            continue

        status = txn["transaction_status"]

        priority = STATUS_PRIORITY.get(status, 0)

        amount = float(txn["total_amount"] or 0)

        current = tx_shipping_candidates.get(order_id)

        # New / higher-priority lifecycle
        if current is None or priority > current["priority"]:
            tx_shipping_candidates[order_id] = {
                "priority": priority,
                "amount": amount,
                "status": status,
            }

        # Same lifecycle → accumulate
        elif priority == current["priority"]:
            current["amount"] += amount

    # ------------------------------------------------------------
    # AFN / FBA SHIPPING
    #
    # Shipment
    #     ↓
    # FBAWeightBasedFee
    # ------------------------------------------------------------

    afn_txns = AmazonTransaction.objects.filter(
        id__in=tx_to_order.keys(),
        transaction_type="Shipment",
        transaction_status__in=[
            "DEFERRED",
            "DEFERRED_RELEASED",
            "RELEASED",
        ],
    ).values(
        "id",
        "transaction_status",
    )

    afn_tx_status = {txn["id"]: txn["transaction_status"] for txn in afn_txns}

    afn_breakdowns = (
        AmazonTransactionBreakdown.objects.filter(
            transaction_id__in=afn_tx_status.keys(),
            breakdown_type="FBAWeightBasedFee",
        )
        .values("transaction_id")
        .annotate(total=Sum("amount"))
    )

    for bd in afn_breakdowns:
        transaction_id = bd["transaction_id"]

        order_id = tx_to_order.get(transaction_id)

        if not order_id:
            continue

        status = afn_tx_status.get(transaction_id)

        if not status:
            continue

        priority = STATUS_PRIORITY.get(status, 0)

        amount = float(bd["total"] or 0)

        current = tx_shipping_candidates.get(order_id)

        # New / higher-priority lifecycle
        if current is None or priority > current["priority"]:
            tx_shipping_candidates[order_id] = {
                "priority": priority,
                "amount": amount,
                "status": status,
            }

        # Same lifecycle → accumulate
        elif priority == current["priority"]:
            current["amount"] += amount

    # ------------------------------------------------------------
    # FINAL SHIPPING MAP
    # ------------------------------------------------------------

    tx_shipping_map = {
        order_id: data["amount"] for order_id, data in tx_shipping_candidates.items()
    }

    print("tx_shipping_map", tx_shipping_map)
    
    
    # ============================================================
    # RETURN CLASSIFICATION (COURIER vs CUSTOMER) — matched by order_id
    # ============================================================
    FULFILLMENT_FEE_REFUND_PATTERNS = ["FulfillmentFeeRefund"]

    refund_txns = AmazonTransaction.objects.filter(
        amazon_account__user=user,
        transaction_type='Refund',
        transaction_status__in=['DEFERRED', 'DEFERRED_RELEASED']
    )

    refund_identifiers = AmazonTransactionRelatedIdentifier.objects.filter(
        transaction__in=refund_txns,
        identifier_name='ORDER_ID',
        identifier_value__in=matching_order_ids
    ).values('transaction_id', 'identifier_value')

    refund_tx_to_order = {
        row['transaction_id']: row['identifier_value']
        for row in refund_identifiers
    }

    refunded_sales_breakdowns = (
        AmazonTransactionBreakdown.objects.filter(
            transaction_id__in=refund_tx_to_order.keys(),
            breakdown_type="Refunded Sales",
        )
        .values("transaction_id")
        .annotate(total=Sum("amount"))
    )
    
    refunded_sales_by_order = {}
    
    for row in refunded_sales_breakdowns:
        order_id = refund_tx_to_order.get(row["transaction_id"])
        if not order_id:
            continue
    
        refunded_sales_by_order[order_id] = (
            refunded_sales_by_order.get(order_id, 0.0)
            + float(row["total"] or 0)
        )
        

    order_ids_with_refund = set(refund_tx_to_order.values())

    fee_refund_q = Q()
    for pattern in FULFILLMENT_FEE_REFUND_PATTERNS:
        fee_refund_q |= Q(description__icontains=pattern)

    fee_refund_txns = AmazonTransaction.objects.filter(
        amazon_account__user=user,
        transaction_type='ServiceFee',
        transaction_status__in=['DEFERRED', 'DEFERRED_RELEASED'],
    ).filter(fee_refund_q)

    fee_refund_identifiers = AmazonTransactionRelatedIdentifier.objects.filter(
        transaction__in=fee_refund_txns,
        identifier_name='ORDER_ID',
        identifier_value__in=matching_order_ids
    ).values_list('identifier_value', flat=True)

    order_ids_with_fee_refund = set(fee_refund_identifiers)

    # ============================================================
    # FULFILLMENT FEE REFUND MAP
    # ============================================================
    
    fulfillment_fee_refund_breakdowns = (
        AmazonTransaction.objects.filter(
            id__in=tx_to_order.keys(),
            transaction_type="ServiceFee",
            transaction_status__in=["DEFERRED", "DEFERRED_RELEASED"],
            description__icontains="EasyshipFulfillmentFeeRefund",
        )
        .values("id", "total_amount")
    )
    
    # ============================================================
    # AMAZON FEES REFUND MAP
    # ============================================================
    
    amazon_fee_breakdowns = (
        AmazonTransactionBreakdown.objects.filter(
            transaction_id__in=refund_tx_to_order.keys(),
            breakdown_type="AmazonFees",
        )
        .values("transaction_id")
        .annotate(total=Sum("amount"))
    )
    
    amazon_fee_refund_by_order = {}
    
    for row in amazon_fee_breakdowns:
        order_id = refund_tx_to_order.get(row["transaction_id"])
        if not order_id:
            continue
    
        amazon_fee_refund_by_order[order_id] = (
            amazon_fee_refund_by_order.get(order_id, 0.0)
            + float(row["total"] or 0)
        )
    
    fulfillment_fee_refund_by_order = {}
    
    for txn in fulfillment_fee_refund_breakdowns:
        order_id = tx_to_order.get(txn["id"])
        if not order_id:
            continue
    
        fulfillment_fee_refund_by_order[order_id] = (
            fulfillment_fee_refund_by_order.get(order_id, 0.0)
            + float(txn["total_amount"] or 0)
        )

    refund_amount_by_order = {}
    refund_count_by_order = {}
    for txn in refund_txns.filter(id__in=refund_tx_to_order.keys()):
        oid = refund_tx_to_order.get(txn.id)
        if not oid:
            continue
        refund_amount_by_order[oid] = (
            refund_amount_by_order.get(oid, 0.0) + float(txn.total_amount or 0)
        )
        refund_count_by_order[oid] = refund_count_by_order.get(oid, 0) + 1

    courier_return_count = 0
    customer_return_count = 0
    courier_return_price = 0.0
    customer_return_price = 0.0

    for order_id in order_ids_with_refund:
        amount = refund_amount_by_order.get(order_id, 0.0)
        if order_id in order_ids_with_fee_refund:
            courier_return_count += 1
            courier_return_price += amount
        else:
            customer_return_count += 1
            customer_return_price += amount

    total_return_count = courier_return_count + customer_return_count

    # ============================================================
    # CLAIM AMOUNT — Transaction Type "Adjustment", description "SERRACReimbursement"
    # ============================================================
    claim_txns = AmazonTransaction.objects.filter(
        amazon_account__user=user,
        transaction_type='Adjustment',
        description__icontains='SERRACReimbursement',
    )

    claim_identifiers = AmazonTransactionRelatedIdentifier.objects.filter(
        transaction__in=claim_txns,
        identifier_name='ORDER_ID',
        identifier_value__in=matching_order_ids
    ).values('transaction_id', 'identifier_value')

    claim_tx_to_order = {
        row['transaction_id']: row['identifier_value']
        for row in claim_identifiers
    }

    claim_amount_by_order = {}
    claim_count_by_order = {}
    for txn in claim_txns.filter(id__in=claim_tx_to_order.keys()):
        oid = claim_tx_to_order.get(txn.id)
        if not oid:
            continue
        claim_amount_by_order[oid] = (
            claim_amount_by_order.get(oid, 0.0) + float(txn.total_amount or 0)
        )
        claim_count_by_order[oid] = claim_count_by_order.get(oid, 0) + 1

    total_claim_amount = sum(claim_amount_by_order.values())
    total_claim_count = len(claim_amount_by_order)
    
    
    # ============================================================
    # REPLACEMENT RETURN — Transaction Type "Shipment",
    # description "Order Payment", total_amount = 0
    # ============================================================
    replacement_txns = AmazonTransaction.objects.filter(
        amazon_account__user=user,
        transaction_type='Shipment',
        description='Order Payment',
        total_amount=0,
    )

    replacement_identifiers = AmazonTransactionRelatedIdentifier.objects.filter(
        transaction__in=replacement_txns,
        identifier_name='ORDER_ID',
        identifier_value__in=matching_order_ids
    ).values('transaction_id', 'identifier_value')

    replacement_tx_to_order = {
        row['transaction_id']: row['identifier_value']
        for row in replacement_identifiers
    }

    order_ids_with_replacement = set(replacement_tx_to_order.values())

    replacement_count_by_order = {}
    for txn in replacement_txns.filter(id__in=replacement_tx_to_order.keys()):
        oid = replacement_tx_to_order.get(txn.id)
        if not oid:
            continue
        replacement_count_by_order[oid] = replacement_count_by_order.get(oid, 0) + 1

    total_replacement_return_count = len(order_ids_with_replacement)
    # ---------------- BUILD RESPONSE ----------------
    results = []

    total_sales = total_profit = total_ads = 0
    total_mpfees = total_net_sales = total_qty = total_final_net_qty = 0
    total_final_net_sales = 0
    total_returns = total_shipping = 0
    total_stdcost = 0
    total_ret_percent = 0
    adjusted_gross_sales = 0
    total_estimatefees = 0 
    total_mp_gst = 0
    total_gst = 0
    total_tcs = 0
    total_tds = 0
    total_taxable_value = 0
    total_gst_payable = 0
    total_exp_settlement = 0
    total_promo_discount = 0

    sku_asin_map = {
        normalize_sku(k): v
        for k, v in OrderItem.objects
            .filter(order_filter)
            .values_list('seller_sku', 'asin')
    }

    child_parent_map = {
        row['asin']: (row['parent_asin'] or row['asin'])
        for row in OrderItem.objects
            .filter(order_filter)
            .values('asin', 'parent_asin')
    }

    # ====== PRE-COMPUTE ADS SPEND ======
    from amazon_auth.models import ProductMapping
    
    # ads_metrics_qs = ProductAdMetric.objects.filter(
    #     product_ad__amazon_account__user=user,
    #     product_ad__amazon_account__is_primary=True,
    # )
    # if from_date:
    #     ads_metrics_qs = ads_metrics_qs.filter(report_date__gte=from_date.date())
    # if to_date:
    #     ads_metrics_qs = ads_metrics_qs.filter(report_date__lte=to_date.date())
    
    ads_metrics_qs = ProductAdMetric.objects.filter(
        product_ad__amazon_account__user=user,
        product_ad__amazon_account__is_primary=True,
    )
    if from_date_ist:
        ads_metrics_qs = ads_metrics_qs.filter(report_date__gte=from_date_ist.date())
    if to_date_ist:
        ads_metrics_qs = ads_metrics_qs.filter(report_date__lte=to_date_ist.date())
        
    ads_agg = ads_metrics_qs.values("product_ad__sku").annotate(
        total_ads_cost=Sum("cost"),
        total_ads_sales=Sum("sales"),
        total_ads_clicks=Sum("clicks"),
        total_ads_orders=Sum("orders"),
        total_ads_impressions=Sum("impressions"),
    )
    
    skus_with_ads = [x["product_ad__sku"] for x in ads_agg if x["product_ad__sku"]]
    
    pm_mappings = ProductMapping.objects.filter(account__user=user, seller_sku__in=skus_with_ads).values("seller_sku", "parent_asin", "asin", "product_name", "image_url")
    pm_dict = {m["seller_sku"]: m for m in pm_mappings}
    
    missing_skus = [sku for sku in skus_with_ads if sku not in pm_dict]
    if missing_skus:
        ali_mappings = AmazonListingItem.objects.filter(user=user, sku__in=missing_skus).values("sku", "asin", "item_name", "image_url")
        for ali in ali_mappings:
            if ali["sku"] not in pm_dict:
                pm_dict[ali["sku"]] = {
                    "seller_sku": ali["sku"],
                    "parent_asin": ali["asin"], # Listing items don't explicitly have parent_asin, use asin
                    "asin": ali["asin"],
                    "product_name": ali["item_name"],
                    "image_url": ali["image_url"],
                }

    missing_skus = [sku for sku in skus_with_ads if sku not in pm_dict]
    if missing_skus:
        oi_mappings = OrderItem.objects.filter(order__user=user, seller_sku__in=missing_skus).values("seller_sku", "parent_asin", "asin", "title", "image_url")
        for oi in oi_mappings:
            if oi["seller_sku"] not in pm_dict:
                pm_dict[oi["seller_sku"]] = {
                    "seller_sku": oi["seller_sku"],
                    "parent_asin": oi["parent_asin"],
                    "asin": oi["asin"],
                    "product_name": oi["title"],
                    "image_url": oi["image_url"],
                }
    
    ads_by_parent = {}
    for agg in ads_agg:
        sku = agg["product_ad__sku"]
        if not sku: continue
        
        pm = pm_dict.get(sku, {})
        p_asin = pm.get("parent_asin") or pm.get("asin") or sku
        
        if p_asin not in ads_by_parent:
            ads_by_parent[p_asin] = {
                "title": pm.get("product_name") or p_asin,
                "image_url": pm.get("image_url") or "",
                "cost": 0, "sales": 0, "clicks": 0, "orders": 0, "impressions": 0
            }
        
        ads_by_parent[p_asin]["cost"] += float(agg["total_ads_cost"] or 0)
        ads_by_parent[p_asin]["sales"] += float(agg["total_ads_sales"] or 0)
        ads_by_parent[p_asin]["clicks"] += int(agg["total_ads_clicks"] or 0)
        ads_by_parent[p_asin]["orders"] += int(agg["total_ads_orders"] or 0)
        ads_by_parent[p_asin]["impressions"] += int(agg["total_ads_impressions"] or 0)
    # ===================================

    processed_parent_asins = set()

    from amazon_auth.other_expence import calculate_other_expenses_map
    from_date_local = from_date_ist.date() if from_date_ist else (start_date.date() if start_date else None)
    to_date_local = to_date_ist.date() if to_date_ist else (end_date.date() if end_date else None)

    parent_sku_map = {}
    
    pm_qs = ProductMapping.objects.filter(account__user=user).values('parent_asin', 'asin', 'seller_sku')
    for pm in pm_qs:
        p = pm.get('parent_asin') or pm.get('asin')
        s = pm.get('seller_sku')
        if p and s:
            parent_sku_map.setdefault(p, set()).add(s)

    all_oi_qs = OrderItem.objects.filter(order__user=user).values('parent_asin', 'asin', 'seller_sku').distinct()
    for oi in all_oi_qs:
        p = oi.get('parent_asin') or oi.get('asin')
        s = oi.get('seller_sku')
        if p and s:
            parent_sku_map.setdefault(p, set()).add(s)

    sku_to_parent = {}
    for p, skus in parent_sku_map.items():
        for s in skus:
            sku_to_parent[s] = p

    ali_qs = AmazonListingItem.objects.filter(user=user).values('asin', 'sku')
    for ali in ali_qs:
        s = ali.get('sku')
        a = ali.get('asin')
        p = sku_to_parent.get(s) or a
        if p and s:
            parent_sku_map.setdefault(p, set()).add(s)

    expense_items = []
    for idx, r in enumerate(items):
        g_qty = float(r.get('grossqty') or 0)
        n_qty = max(g_qty, 0)
        f_sales = float(str(r.get('grosssales') or 0))

        p_asin = r.get('parent_asin') or r.get('asin')
        sku_cnt = len(parent_sku_map.get(p_asin, set())) or 1

        expense_items.append({
            'key': idx,
            'marketplace': r.get('channel') or r.get('marketplace') or 'Amazon-India',
            'units': float(n_qty),
            'net_sales': float(f_sales),
            'sku_count': sku_cnt,
            'order_count_for_sku': 1
        })

    other_expenses_map = calculate_other_expenses_map(user, from_date_local, to_date_local, expense_items)

    for idx, row in enumerate(items):
        row_other_expense = float(other_expenses_map.get(idx, 0))
        # asin = row['asin']
        parent_asin = row['parent_asin']
        processed_parent_asins.add(parent_asin)
        # estimated_fees = estimated_fee_map.get(parent_asin, 0)

        fee_data = estimated_fee_map.get(parent_asin, {})

        estimated_fees = fee_data.get("estimated_fees", 0)

        referral_fee = fee_data.get("referral_fee", 0)
        closing_fee = fee_data.get("closing_fee", 0)
        per_item_fee = fee_data.get("per_item_fee", 0)

        fba_fee = fee_data.get("fba_fee", 0)
        fba_pick_pack_fee = fee_data.get("fba_pick_pack_fee", 0)
        fba_weight_handling_fee = fee_data.get("fba_weight_handling_fee", 0)

        tax_amount = fee_data.get("tax_amount", 0)

        gross_qty = int(row['grossqty'] or 0)
        quantity_shipped = int(row['quantity_shipped'] or 0)

        gross_sales = float(str(row['grosssales'] or 0))
        item_tax = float(str(row.get('item_tax') or 0))
        promo_discount = float(str(row.get('promotion_discount') or 0))

        gst_rate = float(str(row.get("sku_gst_rate") or 0))
        tcs_rate = float(str(row.get("sku_tcs_rate") or 0))
        tds_rate = float(str(row.get("sku_tds_rate") or 0))
        standard_cost = float(str(row.get("sku_standard_cost") or 0))

        orders = asin_map.get(parent_asin, [])

        tx_shipping_final = 0.0
        amazon_fee_refund_total = 0.0
        fulfillment_fee_refund_total = 0.0
        refunded_sales_total = 0.0
        
        for o in orders:
            oid = o['order__amazon_order_id']
            tx_shipping_final += float(tx_shipping_map.get(oid, 0.0))
            amazon_fee_refund_total += float(amazon_fee_refund_by_order.get(oid, 0.0))
            fulfillment_fee_refund_total += float(fulfillment_fee_refund_by_order.get(oid, 0.0))
            refunded_sales_total += float(refunded_sales_by_order.get(oid, 0.0))
            
        shipping_price = tx_shipping_final
        estimated_fees -= amazon_fee_refund_total

        # ==========================================================
        # ADS SPEND (FROM PRE-COMPUTED GLOBALLY)
        # ==========================================================

        parent_ad_data = ads_by_parent.get(parent_asin, {})
        ads = -abs(float(parent_ad_data.get("cost", 0)))
        ads_sales = float(parent_ad_data.get("sales", 0))
        ads_clicks = int(parent_ad_data.get("clicks", 0))
        ads_orders = int(parent_ad_data.get("orders", 0))
        ads_impressions = int(parent_ad_data.get("impressions", 0))

        refund = rto = mpfees = shipping_fee = 0.0
        return_units = 0.0
        t_new_charge = 0.0
        gst = 0.0

        final_net_sales = 0.0
        total_cost = 0.0

        for o in orders:
            oid = o['order__amazon_order_id']
            qty = float(o['quantity_ordered'] or 0)
            o_item_price = float(str(o.get('item_price') or 0))
            o_new_item_price = float(str(o.get('new_item_price') or 0)) 
            o_item_tax = float(str(o.get('item_tax') or 0))

            f = finance_map.get(oid, {})

            refund += float(f.get('refund') or 0)
            rto += float(f.get('rto') or 0)

            mpfees += (
                float(f.get('commission') or 0) +
                float(f.get('fulfillment') or 0) +
                float(f.get('other_fee') or 0)
            )

            shipping_fee += float(f.get('shipping_fee') or 0)
            gst += float(f.get('gst') or 0)

            order_fee_map = extract_fees_and_tcs_per_asin(
                raw_data_map.get(oid, []),
                sku_asin_map=sku_asin_map
            )

            for child_asin, fee_data_inner in order_fee_map.items():
                parent_key = child_parent_map.get(child_asin)
                if parent_key == parent_asin:
                    t_new_charge += float(fee_data_inner["fee"])

            r = float(f.get('refund') or 0)
            rto_amt = float(f.get('rto') or 0)

            if r < 0 or rto_amt < 0:
                return_units += qty

            # Calculate final net sales and cost for this specific order
            
            o_item_price = (
                o_new_item_price
                if o_item_price == 0
                else o_item_price
            )
            o_gross = o_item_price + o_item_tax
            
            print("o_new_item_price newwwwwwwww>>>>>>>>>>>>>>>>",o_new_item_price)
            
            print("o_item_price first>>>>>>>>>>>>>>>>",o_item_price)
            
            print("o_gross first>>>>>>>>>>>>>>>>",o_gross)
            
            # o_gross = o_item_price + o_item_tax
            o_cost = standard_cost * qty

            o_replacement_count = replacement_count_by_order.get(oid, 0)
            o_return_count = refund_count_by_order.get(oid, 0)
            o_has_return = oid in order_ids_with_refund

            if o_replacement_count or (o_has_return and qty == o_return_count):
                o_gross = 0.0
                o_cost = 0.0
                o_promo = float(str(o.get('promotion_discount') or 0))
                promo_discount -= o_promo

            final_net_sales += o_gross
            total_cost += o_cost

        # ------------------------------------------------------------
        # RETURN / CLAIM — aggregated across all orders for this parent_asin row
        # ------------------------------------------------------------
        row_order_ids = [o['order__amazon_order_id'] for o in orders]
        order_fulfillment_fee_refund = sum(
            fulfillment_fee_refund_by_order.get(oid, 0.0) for oid in row_order_ids
        )
            
        order_return_amount = sum(refund_amount_by_order.get(oid, 0.0) for oid in row_order_ids)
        order_refunded_sales = sum(
            refunded_sales_by_order.get(oid, 0.0) for oid in row_order_ids
        )
        order_return_count = sum(refund_count_by_order.get(oid, 0) for oid in row_order_ids)
        order_has_return = any(oid in order_ids_with_refund for oid in row_order_ids)
        order_is_courier_return = any(oid in order_ids_with_fee_refund for oid in row_order_ids)

        if order_has_return and order_is_courier_return:
            order_return_type = "COURIER_RETURN"
        elif order_has_return:
            order_return_type = "CUSTOMER_RETURN"
        else:
            order_return_type = None

        row_courier_return_count = 0
        row_customer_return_count = 0
        row_courier_return_price = 0.0
        row_customer_return_price = 0.0

        seen_order_ids_for_row = set(oid for oid in row_order_ids if oid in order_ids_with_refund)
        for oid in seen_order_ids_for_row:
            amount = refund_amount_by_order.get(oid, 0.0)
            if oid in order_ids_with_fee_refund:
                row_courier_return_count += 1
                row_courier_return_price += amount
            else:
                row_customer_return_count += 1
                row_customer_return_price += amount

        order_claim_amount = sum(claim_amount_by_order.get(oid, 0.0) for oid in row_order_ids)
        order_claim_count = sum(claim_count_by_order.get(oid, 0) for oid in row_order_ids)
        order_has_claim = order_claim_count > 0 
        
        order_replacement_count = sum(replacement_count_by_order.get(oid, 0) for oid in row_order_ids)
        order_is_replacement = any(oid in order_ids_with_replacement for oid in row_order_ids)

        # ------------------------------------------------------------
        # TAXABLE VALUE
        # ------------------------------------------------------------
        
        if gst_rate > 0:
            taxable_value = (
                final_net_sales / (1 + (gst_rate / 100.0))
            )
            gst_to_pay_amount = final_net_sales - taxable_value
        else:
            taxable_value = final_net_sales
            gst_to_pay_amount = 0.0

        # ------------------------------------------------------------
        # TCS  GST TO PAY
        # ------------------------------------------------------------

        if tcs_rate:
            tcs_total = taxable_value * (tcs_rate / 100.0)
        else:
            tcs_total = taxable_value * 0.01

        if tds_rate:
            tds_total = taxable_value * (tds_rate / 100.0)
        else:
            tds_total = 0.0

        if gst_rate:
            gst_to_pay_perc = gst_rate
        else:
            gst_to_pay_perc = (
                (gst_to_pay_amount / taxable_value) * 100.0
                if taxable_value else 0.0
            )  

        # ---------------- CALCULATIONS ----------------
        net_qty = max(gross_qty , 0)
        
        final_net_qty = max(gross_qty , 0)
    
        net_sales = gross_sales + item_tax
        
        shipping_final = ( shipping_price + order_fulfillment_fee_refund ) 

        mp_gst = (-abs(estimated_fees) + shipping_final) * 0.18

        stdcost = total_cost
        stdcost_per_unit = (total_cost / gross_qty) if gross_qty else 0

        avg_cost = float(row.get('avg_cost') or 0)
        missing_qty = 0
        for o in orders:
            if o.get('quantity_ordered') and avg_cost == 0:
                missing_qty += o['quantity_ordered']

        stdcost_missing_percentage = (missing_qty / gross_qty * 100) if gross_qty else 0
        
        # exp_settlement = (
        #     final_net_sales
        #     + shipping_final
        #     + ads
        #     + tcs_total
        #     - estimated_fees
        #     - mp_gst
        #     - promo_discount
        #     - order_claim_amount
        # )
        
        exp_settlement = (
            final_net_sales
            + shipping_final
            # + ads                        remove this 
            - tcs_total                    #substract now 
            - tds_total                    #substract now 
            - estimated_fees
            - mp_gst
            - promo_discount
            + order_claim_amount            #add this one 
        )
        
        profit = (
            final_net_sales
            + shipping_final
            + (ads if profit_setting.ad_spend else 0)
            + (tcs_total if profit_setting.tcs else 0)
            + (tds_total if profit_setting.tds else 0)
            - estimated_fees
            - (mp_gst if profit_setting.input_gst_itc else 0)
            - (gst_to_pay_amount if profit_setting.output_gst else 0)
            - promo_discount
            - (order_claim_amount if profit_setting.claim else 0)
            - (stdcost if profit_setting.product_cost else 0)
            - (row_other_expense if profit_setting.other_expense else 0)
        )
        profit_margin = (profit / final_net_sales * 100) if final_net_sales else 0

        tacos = (
            abs(ads) / gross_sales * 100
        ) if gross_sales else 0
        
        row_customer_return_count += order_replacement_count
        order_return_count += order_replacement_count
        final_net_qty = final_net_qty - order_return_count        
        
        ret_percent = (order_return_count / final_net_qty * 100) if final_net_qty else 0
    
        sku_results.append({
            # "asin": asin,
            "asin": parent_asin,
            "sku": parent_asin, 
            "parent_asin": parent_asin, 
            "name": row['title'],
            "image_url": row['image_url'],
            "channel": "Amazon-India",
            "channel1": "Amazon-India",
            "grossqty": gross_qty,
            "netqty": net_qty,
            "final_net_qty":final_net_qty,   # final_net_qty - all retur
            "grosssales": float(gross_sales),
            "netsales": float(net_sales),
            "final_net_sales": float(final_net_sales),
            # "ads": float(ads),
            "ads": float(ads),
            "ads_sales": float(ads_sales),
            "ads_clicks": ads_clicks,
            "ads_orders": ads_orders,
            "ads_impressions": ads_impressions,
            "mpfees": round(mpfees, 2),
            "mp_gst": float(mp_gst),
            "new_mpfees": float(t_new_charge),
            # "estimatefees": float(estimated_fees),
            "estimatefees": float(-abs(estimated_fees)),

            "referral_fee": float(referral_fee),
            "closing_fee": float(closing_fee),
            "per_item_fee": float(per_item_fee),

            "fba_fee": float(fba_fee),
            "fba_pick_pack_fee": float(fba_pick_pack_fee),
            "fba_weight_handling_fee": float(fba_weight_handling_fee),

            "tax_amount": float(tax_amount),
            "shippingfees": float(shipping_final),
            "profit": float(profit),
            "grossprofitper": round(profit_margin, 2),
            "returnqty": order_return_count,
            "retpercent": round(ret_percent, 2),
            "tacos": round(tacos, 2),
            # "id": asin,
            "id": parent_asin,
            "stdcost": float(stdcost),
            "stdcost_per_unit": round(stdcost_per_unit, 2),
            "stdcostmissingqty": missing_qty,
            "stdcost_missing_percentage": round(stdcost_missing_percentage, 2),
            "redirecturl": f"https://www.amazon.in/dp/{parent_asin}" if parent_asin else None,
            "gst": float(0),
            # "gst": "0",
            "tcs": float(tcs_total),
            "tds": float(tds_total),
            "other_expenses": float(row_other_expense),
            "total_other_expenses": float(row_other_expense),
            "cancelled_qty": 0,
            "cancelled_sales": 0.0,
            "taxable_value": float(taxable_value),
            "gst_to_pay_amount": float(gst_to_pay_amount),
            "gst_to_pay_perc": round(gst_to_pay_perc, 2),
            "exp_settlement": float(exp_settlement),
            
            "promo_discount": float(promo_discount),

            "return_type": order_return_type,
            "is_return": order_has_return,
            "return_count": order_return_count,
            "return_amount": float(order_return_amount),
            
            "courier_return_count": row_courier_return_count,
            "customer_return_count": row_customer_return_count,
            "courier_return_price": float(row_courier_return_price),
            "customer_return_price": float(row_customer_return_price),

            "is_claim": order_has_claim,
            "claim_count": order_claim_count,
            "claim_amount": float(order_claim_amount),
            
            "is_replacement_return": order_is_replacement,
            "replacement_return_count": order_replacement_count,
        })

        # -------- TOTALS --------
        total_sales += gross_sales
        total_net_sales += net_sales
        total_final_net_sales +=  final_net_sales
        total_profit += profit
        total_ads += ads
        total_mpfees += t_new_charge
        total_qty += net_qty
        total_final_net_qty += final_net_qty
        total_returns += return_units
        total_shipping += shipping_final
        total_stdcost += stdcost
        total_gst += gst
        total_tcs += tcs_total
        total_tds += tds_total
        
        total_estimatefees += estimated_fees
        total_mp_gst += mp_gst

        total_taxable_value += taxable_value
        total_gst_payable += gst_to_pay_amount
        total_exp_settlement += exp_settlement
        total_promo_discount += promo_discount
        
        total_return_count += order_replacement_count
        
        customer_return_count += order_replacement_count
        total_ret_percent = (total_return_count / total_final_net_qty * 100) if total_final_net_qty else 0
    # ====== START: ADD ASINS WITH AD SPEND BUT NO ORDERS ======
    for p_asin, data in ads_by_parent.items():
        if p_asin in processed_parent_asins:
            continue
            
        if parent_ids and p_asin not in parent_ids:
            continue
            
        ads_cost = -abs(data["cost"])
        if ads_cost == 0:
            continue
        ads_margin = (ads_cost / 100 * 100) if 1 else 0
        # ads_margin = 0
        sku_results.append({
            "asin": p_asin, 
            "parent_asin": p_asin, 
            "name": data["title"],
            "image_url": data["image_url"],
            "channel": "Amazon-India",
            "channel1": "Amazon-India",
            "grossqty": 0,
            "netqty": 0,
            "final_net_qty": 0,
            "grosssales": float(0),
            "netsales": float(0),
            "ads": float(ads_cost),
            "ads_sales": float(data["sales"]),
            "ads_clicks": data["clicks"],
            "ads_orders": data["orders"],
            "ads_impressions": data["impressions"],
            "mpfees": 0,
            "mp_gst": float(0),
            "new_mpfees": float(0),
            "estimatefees": float(0),
            "referral_fee": float(0),
            "closing_fee": float(0),
            "per_item_fee": float(0),
            "fba_fee": float(0),
            "fba_pick_pack_fee": float(0),
            "fba_weight_handling_fee": float(0),
            "tax_amount": float(0),
            "shippingfees": float(0),
            "profit": float(ads_cost),
            "grossprofitper": round(ads_margin, 2),
            "returnqty": 0,
            "retpercent": 0,
            "tacos": 0,
            "id": p_asin,
            "stdcost": float(0),
            "stdcost_per_unit": 0,
            "stdcostmissingqty": 0,
            "stdcost_missing_percentage": 0,
            "redirecturl": f"https://www.amazon.in/dp/{p_asin}" if p_asin else None,
            "gst": float(0),
            "tcs": float(0),
            "tds": float(0),
            "taxable_value": float(0),
            "gst_to_pay_amount": float(0),
            "gst_to_pay_perc": 0,
            "exp_settlement": float(0),
            "promo_discount": float(0),
            "return_type": None,
            "is_return": False,
            "return_count": 0,
            "return_amount": float(0),
            "courier_return_count": 0,
            "customer_return_count": 0,
            "courier_return_price": float(0),
            "customer_return_price": float(0),
            "is_claim": False,
            "claim_count": 0,
            "claim_amount": float(0),
            "is_replacement_return": False,
            "replacement_return_count": 0,
        })
        
        total_ads += ads_cost
        total_profit += ads_cost

    

    return_claim_summary = {
        "total_return_count": total_return_count,
        "courier_return_count": courier_return_count,
        "customer_return_count": customer_return_count,
        "total_return_amount": float(courier_return_price + customer_return_price),
        "courier_return_amount": float(courier_return_price),
        "customer_return_amount": float(customer_return_price),
        "total_claim_count": total_claim_count,
        "total_claim_amount": float(total_claim_amount),
        "replacement_return_count": total_replacement_return_count,
    }
    return sku_results, return_claim_summary



def export_order_to_excel(file_path="orders.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    fields = [field.name for field in Order._meta.fields]
    ws.append(fields)

    for obj in Order.objects.iterator(chunk_size=1000):
        row = []
        for field in fields:
            value = getattr(obj, field)

            if value is not None:
                value = str(value)
            else:
                value = ""

            row.append(value)

        ws.append(row)

    wb.save(file_path)


def export_all_tables_to_excel(file_path="all_data.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    for model in apps.get_models():
        sheet = wb.create_sheet(title=model.__name__)

        fields = [field.name for field in model._meta.fields]
        sheet.append(fields)

        for obj in model.objects.iterator(chunk_size=1000):
            sheet.append([
                str(getattr(obj, f)) if getattr(obj, f) else ""
                for f in fields
            ])

    wb.save(file_path)


def export_order_items_to_excel(file_path="order_items.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "OrderItems"

    fields = [field.name for field in OrderItem._meta.fields]
    ws.append(fields)

    for obj in OrderItem.objects.iterator(chunk_size=1000):
        row = []
        for field in fields:
            value = getattr(obj, field)

            # 🔹 Handle Foreign Keys nicely
            if field == "order":
                value = obj.order.id if obj.order else ""
            elif field == "product":
                value = str(obj.product) if hasattr(obj, "product") and obj.product else ""

            # 🔹 Handle general values
            elif value is not None:
                value = str(value)
            else:
                value = ""

            row.append(value)

        ws.append(row)

    wb.save(file_path)    



def export_financial_events_to_excel(file_path="financial_events.xlsx"):

    wb = Workbook()
    ws = wb.active
    ws.title = "FinancialEvents"

    fields = [field.name for field in FinancialEvent._meta.fields]
    ws.append(fields)

    for obj in FinancialEvent.objects.iterator(chunk_size=1000):
        row = []
        for field in fields:
            value = getattr(obj, field)

            # Better readable values
            if field == "user":
                value = obj.user.username if obj.user else ""
            elif field == "amazon_account":
                value = str(obj.amazon_account) if obj.amazon_account else ""
            elif field == "raw_data":
                value = str(value)[:500] if value else ""  # avoid huge JSON

            else:
                value = str(value) if value is not None else ""

            row.append(value)

        ws.append(row)

    wb.save(file_path)    



def export_all_data_to_single_excel(file_path="complete_data.xlsx"):


    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---------------------------
    # 🔹 Orders Sheet
    # ---------------------------
    ws_orders = wb.create_sheet(title="Orders")
    order_fields = [f.name for f in Order._meta.fields]
    ws_orders.append(order_fields)

    for obj in Order.objects.iterator(chunk_size=1000):
        ws_orders.append([
            str(getattr(obj, f)) if getattr(obj, f) else ""
            for f in order_fields
        ])

    # ---------------------------
    # 🔹 OrderItems Sheet
    # ---------------------------
    ws_items = wb.create_sheet(title="OrderItems")
    item_fields = [f.name for f in OrderItem._meta.fields]
    ws_items.append(item_fields)

    for obj in OrderItem.objects.iterator(chunk_size=1000):
        row = []
        for f in item_fields:
            value = getattr(obj, f)

            if f == "order":
                value = obj.order.id if obj.order else ""
            elif hasattr(value, "__str__"):
                value = str(value)
            else:
                value = value if value else ""

            row.append(value)

        ws_items.append(row)

    # ---------------------------
    # 🔹 Financial Events Sheet
    # ---------------------------
    ws_fin = wb.create_sheet(title="FinancialEvents")
    fin_fields = [f.name for f in FinancialEvent._meta.fields]
    ws_fin.append(fin_fields)

    for obj in FinancialEvent.objects.iterator(chunk_size=1000):
        row = []
        for f in fin_fields:
            value = getattr(obj, f)

            if f == "user":
                value = obj.user.username if obj.user else ""
            elif f == "amazon_account":
                value = str(obj.amazon_account) if obj.amazon_account else ""
            elif f == "raw_data":
                value = str(value)[:300] if value else ""  # limit size
            else:
                value = str(value) if value else ""

            row.append(value)

        ws_fin.append(row)

    # ---------------------------
    # 🔹 Reports Sheet
    # ---------------------------
    ws_rep = wb.create_sheet(title="Reports")
    rep_fields = [f.name for f in Report._meta.fields]
    ws_rep.append(rep_fields)

    for obj in Report.objects.iterator(chunk_size=1000):
        row = []
        for f in rep_fields:
            value = getattr(obj, f)

            if f == "user":
                value = obj.user.username if obj.user else ""
            elif f == "amazon_account":
                value = str(obj.amazon_account) if obj.amazon_account else ""
            elif f == "raw_data":
                value = str(value)[:300] if value else ""
            else:
                value = str(value) if value else ""

            row.append(value)

        ws_rep.append(row)

    # ---------------------------
    #  Save file
    # ---------------------------
    wb.save(file_path)


def update_orderitem_from_mapping():
    mappings = ProductMapping.objects.all()

    for m in mappings:
        OrderItem.objects.filter(seller_sku=m.seller_sku).update(
            parent_sku=m.parent_sku,
            product_name=m.product_name,
            brand=m.brand,
            cost_price=m.cost_price
        )
        

# amazon_auth/services/ad_importer.py


def safe_get(row, *keys):
    for k in keys:
        if k in row and pd.notna(row[k]):
            return row[k]
    return None


def import_ads_from_excel(file_path):
    # df = pd.read_excel(file_path)
    # df = pd.read_excel(file_path, engine="openpyxl")
    if not os.path.exists(file_path):
        raise Exception(f"File not found: {file_path}")

    if os.path.getsize(file_path) == 0:
        raise Exception("Ads file is empty")

    try:
        df = pd.read_excel(file_path, engine="openpyxl")
    except:
        df = pd.read_csv(file_path)

    for _, row in df.iterrows():
        sku = safe_get(row, 'advertised SKU', 'Advertised SKU', 'sku')
        if not sku:
            continue

        date_val = row.get('date')
        if pd.isna(date_val):
            continue

        AdReport.objects.update_or_create(
            sku=sku,
            date=pd.to_datetime(date_val).date(),
            defaults={
                "impressions": int(row.get('impressions', 0) or 0),
                "clicks": int(row.get('clicks', 0) or 0),
                "spend": float(row.get('spend', 0) or 0),
                "ad_sales": float(row.get('7 day total sales', 0) or 0),
                "ad_orders": int(row.get('7 day total orders', 0) or 0),
            }
        )

# from .spapi_manager import get_catalog_item

def safe_catalog_call(manager, asin, marketplace_id, retries=3):
    print("call catlog function")
    for attempt in range(retries):
        try:
            return manager.get_catalog_item(asin, marketplace_id)
        except Exception as e:
            print(f"Catalog API retry {attempt+1}: {e}")
            time.sleep(2 * (attempt + 1))  # exponential backoff
    return {}



def normalize_financial_events(payload):
    data = payload.get("payload", {}).get("FinancialEvents", {})

    result = {
        "shipments": [],
        "refunds": [],
        "fees": [],
        "adjustments": [],
        "summary": {
            "total_sales": 0,
            "total_refunds": 0,
            "total_fees": 0,
            "net": 0
        }
    }

    # ------------------
    # SHIPMENTS
    # ------------------
    for event in data.get("ShipmentEventList", []):
        total = 0

        for charge in event.get("OrderChargeList", []):
            total += charge["ChargeAmount"]["CurrencyAmount"]

        result["shipments"].append({
            "order_id": event.get("AmazonOrderId"),
            "posted_date": event.get("PostedDate"),
            "amount": total
        })

        result["summary"]["total_sales"] += total

    # ------------------
    # REFUNDS
    # ------------------
    for event in data.get("RefundEventList", []):
        total = 0

        for charge in event.get("OrderChargeList", []):
            total += charge["ChargeAmount"]["CurrencyAmount"]

        result["refunds"].append({
            "order_id": event.get("AmazonOrderId"),
            "posted_date": event.get("PostedDate"),
            "amount": total
        })

        result["summary"]["total_refunds"] += total

    # ------------------
    # FEES (from shipments + refunds)
    # ------------------
    for event_list_name in ["ShipmentEventList", "RefundEventList"]:
        for event in data.get(event_list_name, []):
            for fee in event.get("OrderFeeList", []):
                amount = fee["FeeAmount"]["CurrencyAmount"]

                result["fees"].append({
                    "type": fee["FeeType"],
                    "amount": amount
                })

                result["summary"]["total_fees"] += amount

    # ------------------
    # NET CALCULATION
    # ------------------
    result["summary"]["net"] = (
        result["summary"]["total_sales"]
        - result["summary"]["total_refunds"]
        - result["summary"]["total_fees"]
    )

    return result


    

def classify_event(event_type):
    if event_type == "ShipmentEvent":
        return "SALE"
    elif event_type == "RefundEvent":
        return "REFUND"
    elif event_type in ["GuaranteeClaimEvent", "ChargebackEvent"]:
        return "CLAIM"
    elif event_type in ["ServiceFeeEvent", "FeeEvent"]:
        return "FEE"
    elif event_type == "AdjustmentEvent":
        return "ADJUSTMENT"
    
    elif event_type in ["RetrochargeEvent"]:
        return "RTO"   # 🔥 IMPORTANT
    else:
        return "OTHER"

def get_val(row, *keys, default=0):
    for k in keys:
        if k in row and row[k] not in [None, ""]:
            return row[k]
    return default

def format_currency(value):
    if isinstance(value, str):
        cleaned = value.replace("₹", "").replace(",", "").replace(" ", "").strip()
        try:
            value = float(cleaned) if cleaned else 0.0
        except ValueError:
            value = 0.0
    elif isinstance(value, Decimal):
        value = float(value)
    elif value is None:
        value = 0.0
    else:
        try:
            value = float(value)
        except (ValueError, TypeError):
            value = 0.0
    return f"-₹{abs(round(value, 2))}" if value < 0 else f"₹{round(value, 2)}"


from decimal import Decimal


def extract_fees_and_tcs_per_asin(raw_list, sku_asin_map=None):
    asin_map = {}

    for raw in raw_list:
        if not isinstance(raw, dict):
            continue

        item_lists = []
        item_lists.extend(raw.get("ShipmentItemList", []))
        item_lists.extend(raw.get("ShipmentItemAdjustmentList", []))

        for item in item_lists:
            asin = item.get("ASIN")
            sku = normalize_sku(item.get("SellerSKU"))

            if not asin and sku and sku_asin_map:
                asin = sku_asin_map.get(sku)

            if not asin:
                continue

            asin_map.setdefault(asin, {"fee": Decimal(0), "tcs": Decimal(0)})

            for fee in item.get("ItemFeeList", []) + item.get("ItemFeeAdjustmentList", []):
                asin_map[asin]["fee"] += Decimal(
                    fee.get("FeeAmount", {}).get("CurrencyAmount", 0) or 0
                )

            for charge in item.get("ItemChargeList", []):
                if charge.get("ChargeType") == "TCS-IGST":
                    asin_map[asin]["tcs"] += Decimal(
                        charge.get("ChargeAmount", {}).get("CurrencyAmount", 0) or 0
                    )

    return asin_map

def normalize_sku(sku):
        return sku.replace(" COPY", "").strip() if sku else sku

def clean_sku(sku):
    if not sku:
        return sku
    return sku.replace("       ", "").strip()




def parse_ist_range(from_date_str, to_date_str, end_exclusive=True):
    """
    Parses fromDate/toDate strings (YYYY-MM-DD, assumed IST calendar days)
    into both a UTC datetime range (for purchase_date/posted_date filters)
    and a local date range (for report_date / other local DateField filters).

    end_exclusive=True mimics the existing '+1 day' pattern used for
    order__purchase_date__lte comparisons in several endpoints.
    """
    from_date_utc = to_date_utc = None
    from_date_local = to_date_local = None

    if from_date_str:
        from_date_local = datetime.strptime(from_date_str[:10], "%Y-%m-%d").date()
        from_date_utc = datetime.combine(from_date_local, datetime.min.time(), tzinfo=IST).astimezone(UTC)

    if to_date_str:
        to_date_local = datetime.strptime(to_date_str[:10], "%Y-%m-%d").date()
        if end_exclusive:
            naive_to = datetime.combine(to_date_local, datetime.min.time()) + timedelta(days=1)
            to_date_utc = naive_to.replace(tzinfo=IST).astimezone(UTC)
        else:
            to_date_utc = datetime.combine(to_date_local, datetime.max.time().replace(microsecond=0), tzinfo=IST).astimezone(UTC)

    if from_date_utc and not to_date_utc:
        to_date_utc = from_date_utc + timedelta(days=1)
        to_date_local = from_date_local

    return from_date_utc, to_date_utc, from_date_local, to_date_local


def filter_ads_by_local_range(ads_qs, from_date_local, to_date_local):
    if from_date_local:
        ads_qs = ads_qs.filter(report_date__gte=from_date_local)
    if to_date_local:
        ads_qs = ads_qs.filter(report_date__lte=to_date_local)
    return ads_qs

