from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import load_workbook

from blinkit.models import (
    BlinkitProduct,
    BlinkitStorageCharge,
)

DAILY_AGEING_SHEET = "Daily Ageing"
AGING_REVERSAL_SHEET = "Aging Reversal"
UPFRONT_STORAGE_SHEET = "Upfront Storage Charges"


# =====================================================================
# BASIC HELPERS
# =====================================================================


def clean_value(value):
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if not value or value in {
            "-",
            "--",
            "NA",
            "N/A",
        }:
            return None

        return value

    return value


def normalize_text(value):
    value = clean_value(value)

    if value is None:
        return None

    return str(value).strip()


def normalize_id(value):
    """
    Normalize Blinkit IDs.

    Examples:

        10282540
        10282540.0
        "10282540"

    all become:

        "10282540"
    """

    value = clean_value(value)

    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def to_decimal(value, default=Decimal("0")):
    value = clean_value(value)

    if value is None:
        return default

    if isinstance(value, Decimal):
        return value

    try:
        if isinstance(value, str):
            value = value.replace("₹", "").replace(",", "").replace("%", "").strip()

        return Decimal(str(value))

    except (
        InvalidOperation,
        ValueError,
        TypeError,
    ):
        return default


def to_decimal_or_none(value):
    value = clean_value(value)

    if value is None:
        return None

    return to_decimal(value)


def serialize_raw_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()

    if isinstance(value, Decimal):
        return str(value)

    return value


# =====================================================================
# HEADER HELPERS
# =====================================================================


def normalize_header(value):
    value = clean_value(value)

    if value is None:
        return None

    return " ".join(str(value).strip().lower().split())


def build_header_map(worksheet, header_row=4):
    header_map = {}

    for column_number, cell in enumerate(
        worksheet[header_row],
        start=1,
    ):
        header = normalize_header(cell.value)

        if header:
            header_map[header] = column_number

    if not header_map:
        raise ValueError(f"Unable to find headers in sheet '{worksheet.title}'.")

    return header_map


def get_cell_value(
    row,
    header_map,
    column_name,
):
    normalized_column = normalize_header(column_name)

    column_number = header_map.get(normalized_column)

    if column_number is None:
        return None

    index = column_number - 1

    if index >= len(row):
        return None

    return row[index]


def get_storage_rows(
    worksheet,
    data_start_row=5,
):
    for row in worksheet.iter_rows(
        min_row=data_start_row,
        values_only=True,
    ):
        if not any(value is not None for value in row):
            continue

        yield row


# =====================================================================
# STORAGE PERIOD
# =====================================================================


def parse_report_date(value):
    """
    Convert an Excel date/header into Python date.
    """

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        value = value.strip()

        formats = (
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%d-%b-%Y",
            "%d/%b/%Y",
            "%Y-%m-%d %H:%M:%S",
        )

        for fmt in formats:
            try:
                return datetime.strptime(
                    value,
                    fmt,
                ).date()
            except ValueError:
                continue

    return None


def get_storage_period(
    worksheet,
    header_row=4,
):
    """
    Daily Ageing contains one column for each date.

    Example:

        2026-06-01
        2026-06-02
        ...
        2026-06-15

    Therefore:

        period_start = 2026-06-01
        period_end   = 2026-06-15
    """

    dates = []

    for cell in worksheet[header_row]:
        parsed_date = parse_report_date(cell.value)

        if parsed_date:
            dates.append(parsed_date)

    if not dates:
        raise ValueError(
            "Unable to determine storage period from "
            "Daily Ageing report. No date columns were found."
        )

    return min(dates), max(dates)


# =====================================================================
# PRODUCT
# =====================================================================


def find_product(
    *,
    account,
    item_id,
):
    if not item_id:
        return None

    return BlinkitProduct.objects.filter(
        account=account,
        item_id=item_id,
    ).first()


# =====================================================================
# RAW DATA
# =====================================================================


def build_raw_row(
    row,
    worksheet,
):
    """
    Preserve the complete original Excel row.
    """

    headers = []

    for cell in worksheet[4]:
        headers.append(normalize_text(cell.value))

    raw_data = {}

    for index, value in enumerate(row):
        if index >= len(headers):
            continue

        header = headers[index]

        if not header:
            header = f"column_{index + 1}"

        raw_data[header] = serialize_raw_value(value)

    return raw_data


# =====================================================================
# FIND EXISTING STORAGE RECORD
# =====================================================================


def find_existing_storage_charge(
    *,
    account,
    import_batch,
    item_id,
    charge_type,
    ageing_slab=None,
):
    """
    Find the existing storage record for the same:

        account
        payout period
        item_id
        charge_type
        ageing_slab
    """

    queryset = BlinkitStorageCharge.objects.filter(
        account=account,
        source_import__payout_period_start=(import_batch.payout_period_start),
        source_import__payout_period_end=(import_batch.payout_period_end),
        item_id=item_id,
        charge_type=charge_type,
    )

    if charge_type in {
        "DAILY_AGEING",
        "AGING_REVERSAL",
    }:
        queryset = queryset.filter(
            ageing_slab=ageing_slab,
        )

    return queryset.first()


# =====================================================================
# DAILY AGEING
# =====================================================================


def import_daily_ageing_row(
    *,
    row,
    worksheet,
    header_map,
    account,
    import_batch,
):
    item_id = normalize_id(
        get_cell_value(
            row,
            header_map,
            "Item ID",
        )
    )

    if not item_id:
        return {
            "created": False,
            "updated": False,
            "skipped": True,
        }

    item_name = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Item Name",
        )
    )

    state = normalize_text(
        get_cell_value(
            row,
            header_map,
            "State",
        )
    )

    per_day_charge = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Per day charge (Rs)",
        )
    )

    total_inventory_unit_days = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Total inventory unit x days",
        )
    )

    unit_charge = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Unit Charge (Rs)",
        )
    )

    regime = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Regime",
        )
    )

    ageing_slab = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Ageing Slab",
        )
    )

    total_charge = to_decimal(
        get_cell_value(
            row,
            header_map,
            "Total Charge (Rs)",
        )
    )

    raw_data = build_raw_row(
        row,
        worksheet,
    )

    existing = find_existing_storage_charge(
        account=account,
        import_batch=import_batch,
        item_id=item_id,
        charge_type="DAILY_AGEING",
        ageing_slab=ageing_slab,
    )

    product = find_product(
        account=account,
        item_id=item_id,
    )

    if existing:
        changed = False

        fields = {
            "product": product,
            "item_name": item_name,
            "state": state,
            "per_day_charge": per_day_charge,
            "total_inventory_unit_days": (total_inventory_unit_days),
            "unit_charge": unit_charge,
            "regime": regime,
            "ageing_slab": ageing_slab,
            "total_charge": total_charge,
        }

        for field, new_value in fields.items():
            if getattr(existing, field) != new_value:
                setattr(
                    existing,
                    field,
                    new_value,
                )
                changed = True

        if changed:
            existing.raw_data = raw_data
            existing.save()

            return {
                "created": False,
                "updated": True,
                "skipped": False,
                "charge": existing,
            }

        return {
            "created": False,
            "updated": False,
            "skipped": True,
            "charge": existing,
        }

    charge = BlinkitStorageCharge.objects.create(
        account=account,
        product=product,
        source_import=import_batch,
        item_id=item_id,
        item_name=item_name,
        state=state,
        charge_type="DAILY_AGEING",
        per_day_charge=per_day_charge,
        total_inventory_unit_days=(total_inventory_unit_days),
        unit_charge=unit_charge,
        regime=regime,
        ageing_slab=ageing_slab,
        total_charge=total_charge,
        raw_data=raw_data,
    )

    return {
        "created": True,
        "updated": False,
        "skipped": False,
        "charge": charge,
    }


# =====================================================================
# AGING REVERSAL
# =====================================================================


def import_aging_reversal_row(
    *,
    row,
    worksheet,
    header_map,
    account,
    import_batch,
):
    item_id = normalize_id(
        get_cell_value(
            row,
            header_map,
            "Item ID",
        )
    )

    if not item_id:
        return {
            "created": False,
            "updated": False,
            "skipped": True,
        }

    item_name = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Item Name",
        )
    )

    state = normalize_text(
        get_cell_value(
            row,
            header_map,
            "State",
        )
    )

    per_day_charge = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Per day charge (Rs)",
        )
    )

    total_inventory_unit_days = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Total inventory unit x days",
        )
    )

    unit_charge = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Unit Charge (Rs)",
        )
    )

    regime = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Regime",
        )
    )

    ageing_slab = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Ageing Slab",
        )
    )

    total_charge = to_decimal(
        get_cell_value(
            row,
            header_map,
            "Total Charge (Rs)",
        )
    )

    raw_data = build_raw_row(
        row,
        worksheet,
    )

    existing = find_existing_storage_charge(
        account=account,
        import_batch=import_batch,
        item_id=item_id,
        charge_type="AGING_REVERSAL",
        ageing_slab=ageing_slab,
    )

    product = find_product(
        account=account,
        item_id=item_id,
    )

    if existing:
        changed = False

        fields = {
            "product": product,
            "item_name": item_name,
            "state": state,
            "per_day_charge": per_day_charge,
            "total_inventory_unit_days": (total_inventory_unit_days),
            "unit_charge": unit_charge,
            "regime": regime,
            "ageing_slab": ageing_slab,
            "total_charge": total_charge,
        }

        for field, new_value in fields.items():
            if getattr(existing, field) != new_value:
                setattr(
                    existing,
                    field,
                    new_value,
                )
                changed = True

        if changed:
            existing.raw_data = raw_data
            existing.save()

            return {
                "created": False,
                "updated": True,
                "skipped": False,
                "charge": existing,
            }

        return {
            "created": False,
            "updated": False,
            "skipped": True,
            "charge": existing,
        }

    charge = BlinkitStorageCharge.objects.create(
        account=account,
        product=product,
        source_import=import_batch,
        item_id=item_id,
        item_name=item_name,
        state=state,
        charge_type="AGING_REVERSAL",
        per_day_charge=per_day_charge,
        total_inventory_unit_days=(total_inventory_unit_days),
        unit_charge=unit_charge,
        regime=regime,
        ageing_slab=ageing_slab,
        total_charge=total_charge,
        raw_data=raw_data,
    )

    return {
        "created": True,
        "updated": False,
        "skipped": False,
        "charge": charge,
    }


# =====================================================================
# UPFRONT STORAGE
# =====================================================================


def import_upfront_storage_row(
    *,
    row,
    worksheet,
    header_map,
    account,
    import_batch,
):
    item_id = normalize_id(
        get_cell_value(
            row,
            header_map,
            "Item ID",
        )
    )

    if not item_id:
        return {
            "created": False,
            "updated": False,
            "skipped": True,
        }

    item_name = normalize_text(
        get_cell_value(
            row,
            header_map,
            "Item Name",
        )
    )

    state = normalize_text(
        get_cell_value(
            row,
            header_map,
            "State",
        )
    )

    quantity = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Qty",
        )
    )

    unit_charge = to_decimal_or_none(
        get_cell_value(
            row,
            header_map,
            "Unit Charge (Rs)",
        )
    )

    total_charge = to_decimal(
        get_cell_value(
            row,
            header_map,
            "Total Charge (Rs)",
        )
    )

    raw_data = build_raw_row(
        row,
        worksheet,
    )

    existing = find_existing_storage_charge(
        account=account,
        import_batch=import_batch,
        item_id=item_id,
        charge_type="UPFRONT_STORAGE",
    )

    product = find_product(
        account=account,
        item_id=item_id,
    )

    if existing:
        changed = False

        fields = {
            "product": product,
            "item_name": item_name,
            "state": state,
            "quantity": quantity,
            "unit_charge": unit_charge,
            "total_charge": total_charge,
        }

        for field, new_value in fields.items():
            if getattr(existing, field) != new_value:
                setattr(
                    existing,
                    field,
                    new_value,
                )
                changed = True

        if changed:
            existing.raw_data = raw_data
            existing.save()

            return {
                "created": False,
                "updated": True,
                "skipped": False,
                "charge": existing,
            }

        return {
            "created": False,
            "updated": False,
            "skipped": True,
            "charge": existing,
        }

    charge = BlinkitStorageCharge.objects.create(
        account=account,
        product=product,
        source_import=import_batch,
        item_id=item_id,
        item_name=item_name,
        state=state,
        charge_type="UPFRONT_STORAGE",
        quantity=quantity,
        unit_charge=unit_charge,
        total_charge=total_charge,
        raw_data=raw_data,
    )

    return {
        "created": True,
        "updated": False,
        "skipped": False,
        "charge": charge,
    }


# =====================================================================
# MAIN IMPORTER
# =====================================================================


@transaction.atomic
def import_storage_report(
    *,
    file,
    account,
    import_batch,
):
    """
    Import Blinkit Storage Charges workbook.

    Storage period is determined from the date columns
    in Daily Ageing.
    """

    file.seek(0)

    workbook = load_workbook(
        file,
        read_only=True,
        data_only=True,
    )

    # =================================================================
    # Validate sheets
    # =================================================================

    required_sheets = {
        DAILY_AGEING_SHEET,
        AGING_REVERSAL_SHEET,
        UPFRONT_STORAGE_SHEET,
    }

    missing_sheets = required_sheets - set(workbook.sheetnames)

    if missing_sheets:
        raise ValueError(
            "Storage report is missing required sheets: "
            + ", ".join(sorted(missing_sheets))
        )

    # =================================================================
    # Daily Ageing
    # =================================================================

    daily_worksheet = workbook[DAILY_AGEING_SHEET]

    daily_header_map = build_header_map(
        daily_worksheet,
        header_row=4,
    )

    daily_rows = list(
        get_storage_rows(
            daily_worksheet,
            data_start_row=5,
        )
    )

    if not daily_rows:
        raise ValueError("Daily Ageing sheet is empty.")

    # =================================================================
    # Determine storage period
    # =================================================================

    period_start, period_end = get_storage_period(
        daily_worksheet,
        header_row=4,
    )

    import_batch.payout_period_start = period_start
    import_batch.payout_period_end = period_end

    import_batch.save(
        update_fields=[
            "payout_period_start",
            "payout_period_end",
        ]
    )

    # =================================================================
    # Aging Reversal
    # =================================================================

    reversal_worksheet = workbook[AGING_REVERSAL_SHEET]

    reversal_header_map = build_header_map(
        reversal_worksheet,
        header_row=4,
    )

    reversal_rows = list(
        get_storage_rows(
            reversal_worksheet,
            data_start_row=5,
        )
    )

    # =================================================================
    # Upfront Storage
    # =================================================================

    upfront_worksheet = workbook[UPFRONT_STORAGE_SHEET]

    upfront_header_map = build_header_map(
        upfront_worksheet,
        header_row=4,
    )

    upfront_rows = list(
        get_storage_rows(
            upfront_worksheet,
            data_start_row=5,
        )
    )

    total_rows = len(daily_rows) + len(reversal_rows) + len(upfront_rows)

    if total_rows == 0:
        raise ValueError("Blinkit Storage Charges report contains no data.")

    # =================================================================
    # Counters
    # =================================================================

    daily_created = 0
    daily_updated = 0
    daily_skipped = 0

    reversal_created = 0
    reversal_updated = 0
    reversal_skipped = 0

    upfront_created = 0
    upfront_updated = 0
    upfront_skipped = 0

    # =================================================================
    # Daily Ageing
    # =================================================================

    for row in daily_rows:
        result = import_daily_ageing_row(
            row=row,
            worksheet=daily_worksheet,
            header_map=daily_header_map,
            account=account,
            import_batch=import_batch,
        )

        if result["created"]:
            daily_created += 1

        elif result["updated"]:
            daily_updated += 1

        elif result["skipped"]:
            daily_skipped += 1

    # =================================================================
    # Aging Reversal
    # =================================================================

    for row in reversal_rows:
        result = import_aging_reversal_row(
            row=row,
            worksheet=reversal_worksheet,
            header_map=reversal_header_map,
            account=account,
            import_batch=import_batch,
        )

        if result["created"]:
            reversal_created += 1

        elif result["updated"]:
            reversal_updated += 1

        elif result["skipped"]:
            reversal_skipped += 1

    # =================================================================
    # Upfront Storage
    # =================================================================

    for row in upfront_rows:
        result = import_upfront_storage_row(
            row=row,
            worksheet=upfront_worksheet,
            header_map=upfront_header_map,
            account=account,
            import_batch=import_batch,
        )

        if result["created"]:
            upfront_created += 1

        elif result["updated"]:
            upfront_updated += 1

        elif result["skipped"]:
            upfront_skipped += 1

    # =================================================================
    # Result
    # =================================================================

    return {
        "total_rows": total_rows,
        "payout_period_start": str(period_start),
        "payout_period_end": str(period_end),
        "daily_ageing": {
            "rows": len(daily_rows),
            "created": daily_created,
            "updated": daily_updated,
            "skipped": daily_skipped,
        },
        "aging_reversal": {
            "rows": len(reversal_rows),
            "created": reversal_created,
            "updated": reversal_updated,
            "skipped": reversal_skipped,
        },
        "upfront_storage": {
            "rows": len(upfront_rows),
            "created": upfront_created,
            "updated": upfront_updated,
            "skipped": upfront_skipped,
        },
        "total_created": (daily_created + reversal_created + upfront_created),
        "total_updated": (daily_updated + reversal_updated + upfront_updated),
        "total_skipped": (daily_skipped + reversal_skipped + upfront_skipped),
    }
