import io
import openpyxl
from decimal import Decimal
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from rest_framework import serializers, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import MarketplaceEstimatedFeeRule

User = get_user_model()

DEFAULT_MARKETPLACE_TEMPLATES = {
    'Myntra': [
        {
            'name': 'Commission',
            'desc': 'Myntra calls it platform commission',
            'how': 'pct-slab',
            'by_cat': True,
            'on': True,
            'value': 0,
            'groups': [
                {
                    'label': 'Apparel › Tops › Women',
                    'slabs': [
                        [0, 500, 4],
                        [500, 1000, 8],
                        [1000, 2000, 15],
                        [2000, '', 15],
                    ],
                },
                {
                    'label': 'Apparel › Dresses › Women',
                    'slabs': [
                        [0, 800, 4],
                        [800, 2000, 15],
                        [2000, '', 15],
                    ],
                },
            ],
        },
        {
            'name': 'Fixed fee',
            'desc': 'Charged on each item sold',
            'how': 'flat-slab',
            'by_cat': True,
            'on': True,
            'value': 0,
            'groups': [
                {
                    'label': 'Apparel › Tops › Women',
                    'slabs': [
                        [0, 400, 0],
                        [400, 450, 15],
                        [450, 1000, 27],
                        [1000, 2000, 45],
                        [2000, '', 61],
                    ],
                },
                {
                    'label': 'Apparel › Dresses › Women',
                    'slabs': [
                        [0, 500, 0],
                        [500, 600, 3],
                        [600, 1000, 27],
                        [1000, 2000, 45],
                        [2000, '', 61],
                    ],
                },
            ],
        },
        {
            'name': 'Return fee',
            'desc': 'When a customer returns an order',
            'how': 'flat',
            'by_cat': False,
            'on': True,
            'value': 60,
            'groups': [],
        },
        {
            'name': 'Marketing services fee',
            'desc': 'Charged on net sales value',
            'how': 'pct',
            'by_cat': False,
            'on': True,
            'value': 2,
            'groups': [],
        },
        {
            'name': 'Shipping fee',
            'desc': 'Not charged — logistics is in the commission',
            'how': 'flat',
            'by_cat': False,
            'on': False,
            'value': 0,
            'groups': [],
        },
    ]
}


class MarketplaceEstimatedFeeRuleSerializer(serializers.ModelSerializer):
    class Meta:
        model = MarketplaceEstimatedFeeRule
        fields = [
            'id',
            'user',
            'marketplace',
            'name',
            'desc',
            'how',
            'by_cat',
            'on',
            'value',
            'groups',
            'created_at',
            'updated_at',
        ]
        read_only_fields = ['id', 'user', 'created_at', 'updated_at']


class MarketplaceEstimatedFeeRuleListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        marketplace = request.query_params.get('marketplace', 'Myntra')

        qs = MarketplaceEstimatedFeeRule.objects.filter(user=user, marketplace__iexact=marketplace)

        # Seed defaults if user has no rules for this marketplace
        if not qs.exists() and marketplace in DEFAULT_MARKETPLACE_TEMPLATES:
            templates = DEFAULT_MARKETPLACE_TEMPLATES[marketplace]
            created_rules = []
            for tpl in templates:
                rule = MarketplaceEstimatedFeeRule.objects.create(
                    user=user,
                    marketplace=marketplace,
                    name=tpl['name'],
                    desc=tpl['desc'],
                    how=tpl['how'],
                    by_cat=tpl['by_cat'],
                    on=tpl['on'],
                    value=Decimal(str(tpl['value'])),
                    groups=tpl['groups'],
                )
                created_rules.append(rule)
            serializer = MarketplaceEstimatedFeeRuleSerializer(created_rules, many=True)
            return Response({'success': True, 'count': len(created_rules), 'results': serializer.data}, status=status.HTTP_200_OK)

        serializer = MarketplaceEstimatedFeeRuleSerializer(qs, many=True)
        return Response({'success': True, 'count': qs.count(), 'results': serializer.data}, status=status.HTTP_200_OK)

    def post(self, request):
        user = request.user
        serializer = MarketplaceEstimatedFeeRuleSerializer(data=request.data)
        if serializer.is_valid():
            rule = serializer.save(user=user)
            return Response({'success': True, 'message': 'Fee rule created successfully', 'data': MarketplaceEstimatedFeeRuleSerializer(rule).data}, status=status.HTTP_201_CREATED)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)


class MarketplaceEstimatedFeeRuleDetailAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk, user):
        try:
            return MarketplaceEstimatedFeeRule.objects.get(pk=pk, user=user)
        except MarketplaceEstimatedFeeRule.DoesNotExist:
            return None

    def get(self, request, pk):
        rule = self.get_object(pk, request.user)
        if not rule:
            return Response({'success': False, 'message': 'Fee rule not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = MarketplaceEstimatedFeeRuleSerializer(rule)
        return Response({'success': True, 'data': serializer.data}, status=status.HTTP_200_OK)

    def put(self, request, pk):
        rule = self.get_object(pk, request.user)
        if not rule:
            return Response({'success': False, 'message': 'Fee rule not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = MarketplaceEstimatedFeeRuleSerializer(rule, data=request.data, partial=True)
        if serializer.is_valid():
            updated = serializer.save()
            return Response({'success': True, 'message': 'Fee rule updated successfully', 'data': MarketplaceEstimatedFeeRuleSerializer(updated).data}, status=status.HTTP_200_OK)
        return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        rule = self.get_object(pk, request.user)
        if not rule:
            return Response({'success': False, 'message': 'Fee rule not found'}, status=status.HTTP_404_NOT_FOUND)
        rule.delete()
        return Response({'success': True, 'message': 'Fee rule deleted successfully'}, status=status.HTTP_200_OK)


class MarketplaceEstimatedFeeCalculateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        marketplace = request.data.get('marketplace', 'Myntra')
        category = request.data.get('category', "Apparel › Tops › Women")
        try:
            price = Decimal(str(request.data.get('price', '899')))
        except Exception:
            price = Decimal('899.00')

        rules = MarketplaceEstimatedFeeRule.objects.filter(user=user, marketplace__iexact=marketplace, on=True)

        breakdown = []
        total_fees = Decimal('0.00')

        for rule in rules:
            fee_amount = Decimal('0.00')
            applied_rate = None

            if rule.how == 'pct':
                fee_amount = (price * Decimal(str(rule.value))) / Decimal('100')
                applied_rate = f"{rule.value}%"
            elif rule.how == 'flat':
                fee_amount = Decimal(str(rule.value))
                applied_rate = f"₹{rule.value}"
            elif rule.how in ['pct-slab', 'flat-slab']:
                groups = rule.groups or []
                matched_group = None
                if rule.by_cat:
                    matched_group = next((g for g in groups if g.get('label') == category), None)
                if not matched_group and groups:
                    matched_group = groups[0]

                if matched_group:
                    slabs = matched_group.get('slabs', [])
                    for s in slabs:
                        low = Decimal(str(s[0])) if s[0] != '' else Decimal('0')
                        high = Decimal(str(s[1])) if len(s) > 1 and s[1] != '' and s[1] is not None else None
                        rate = Decimal(str(s[2])) if len(s) > 2 and s[2] != '' else Decimal('0')

                        if price >= low and (high is None or price < high):
                            if rule.how == 'pct-slab':
                                fee_amount = (price * rate) / Decimal('100')
                                applied_rate = f"{rate}% slab"
                            else:
                                fee_amount = rate
                                applied_rate = f"₹{rate} slab"
                            break

            total_fees += fee_amount
            breakdown.append({
                'fee_name': rule.name,
                'how': rule.how,
                'applied_rate': applied_rate,
                'amount': float(fee_amount),
            })

        net_amount = price - total_fees

        return Response({
            'success': True,
            'marketplace': marketplace,
            'price': float(price),
            'total_fees': float(total_fees),
            'net_amount': float(net_amount),
            'breakdown': breakdown,
        }, status=status.HTTP_200_OK)


class MarketplaceEstimatedFeeRuleDownloadSampleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rate Card Sample"

        headers = [
            "Marketplace",
            "Fee Name",
            "Description",
            "Calculation Type",
            "Default Value",
            "Product Category",
            "Price From",
            "Price To",
            "Slab Rate",
            "Status",
        ]
        ws.append(headers)

        sample_rows = [
            ["Myntra", "Commission", "Myntra calls it platform commission", "pct-slab", 0, "Apparel › Tops › Women", 0, 500, 4, "On"],
            ["Myntra", "Commission", "Myntra calls it platform commission", "pct-slab", 0, "Apparel › Tops › Women", 500, 1000, 8, "On"],
            ["Myntra", "Commission", "Myntra calls it platform commission", "pct-slab", 0, "Apparel › Tops › Women", 1000, 2000, 15, "On"],
            ["Myntra", "Commission", "Myntra calls it platform commission", "pct-slab", 0, "Apparel › Tops › Women", 2000, "", 15, "On"],
            ["Myntra", "Fixed fee", "Charged on each item sold", "flat-slab", 0, "Apparel › Tops › Women", 0, 400, 0, "On"],
            ["Myntra", "Fixed fee", "Charged on each item sold", "flat-slab", 0, "Apparel › Tops › Women", 400, 450, 15, "On"],
            ["Myntra", "Fixed fee", "Charged on each item sold", "flat-slab", 0, "Apparel › Tops › Women", 450, 1000, 27, "On"],
            ["Myntra", "Return fee", "When a customer returns an order", "flat", 60, "", "", "", "", "On"],
            ["Myntra", "Marketing services fee", "Charged on net sales value", "pct", 2, "", "", "", "", "On"],
            ["Myntra", "Shipping fee", "Not charged", "flat", 0, "", "", "", "", "Off"],
        ]

        for row in sample_rows:
            ws.append(row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = 'attachment; filename="Estimated_Fees_Sample_Template.xlsx"'
        return response


class MarketplaceEstimatedFeeRuleUploadExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        user = request.user
        if 'file' not in request.FILES:
            return Response({'success': False, 'message': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        file_obj = request.FILES['file']
        file_name = file_obj.name.lower()

        try:
            if file_name.endswith('.csv'):
                import csv
                decoded_file = file_obj.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))

            if not rows or len(rows) < 2:
                return Response({'success': False, 'message': 'File is empty or missing data rows'}, status=status.HTTP_400_BAD_REQUEST)

            data_rows = rows[1:]
            rules_dict = {}

            for row in data_rows:
                if not row or not any(row):
                    continue

                marketplace = str(row[0]).strip() if len(row) > 0 and row[0] is not None else 'Myntra'
                fee_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                desc = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                how = str(row[3]).strip().lower() if len(row) > 3 and row[3] is not None else 'pct'
                default_val = float(row[4]) if len(row) > 4 and row[4] not in ['', None] else 0.0
                cat_label = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
                price_from = row[6] if len(row) > 6 and row[6] not in ['', None] else ''
                price_to = row[7] if len(row) > 7 and row[7] not in ['', None] else ''
                slab_rate = row[8] if len(row) > 8 and row[8] not in ['', None] else ''
                status_str = str(row[9]).strip().lower() if len(row) > 9 and row[9] is not None else 'on'
                is_on = status_str in ['on', 'true', '1', 'yes']

                if not fee_name:
                    continue

                key = (marketplace, fee_name)
                if key not in rules_dict:
                    rules_dict[key] = {
                        'marketplace': marketplace,
                        'name': fee_name,
                        'desc': desc,
                        'how': how,
                        'by_cat': bool(cat_label),
                        'on': is_on,
                        'value': Decimal(str(default_val)),
                        'groups_map': {},
                    }

                r_item = rules_dict[key]
                if cat_label:
                    r_item['by_cat'] = True

                if how in ['pct-slab', 'flat-slab'] and (slab_rate != '' or price_from != ''):
                    g_key = cat_label or 'All products'
                    if g_key not in r_item['groups_map']:
                        r_item['groups_map'][g_key] = []

                    s_from = float(price_from) if price_from != '' else 0
                    s_to = float(price_to) if price_to != '' else ''
                    s_rate = float(slab_rate) if slab_rate != '' else 0
                    r_item['groups_map'][g_key].append([s_from, s_to, s_rate])

            processed_count = 0
            for key, r_data in rules_dict.items():
                marketplace, fee_name = key
                groups = []
                for g_label, slabs in r_data['groups_map'].items():
                    groups.append({
                        'label': g_label,
                        'slabs': slabs,
                    })

                MarketplaceEstimatedFeeRule.objects.update_or_create(
                    user=user,
                    marketplace=marketplace,
                    name=fee_name,
                    defaults={
                        'desc': r_data['desc'],
                        'how': r_data['how'],
                        'by_cat': r_data['by_cat'],
                        'on': r_data['on'],
                        'value': r_data['value'],
                        'groups': groups,
                    },
                )
                processed_count += 1

            return Response({
                'success': True,
                'message': f'Successfully imported {processed_count} fee rules from file',
                'count': processed_count,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'success': False, 'message': f'Error parsing file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
